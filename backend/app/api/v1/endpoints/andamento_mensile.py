# backend/app/api/v1/endpoints/andamento_mensile.py
import io
from datetime import date
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.deps import tutti_i_ruoli
from app.models.progetto import Progetto
from app.models.budget import BudgetVoce, VoceDiCosto, Impegno, Spesa
from app.models.autorizzazione_spesa import RichiestaAutorizzazioneSpesa
from app.models.persona import Persona
from app.models.personale import Allocazione
from app.api.v1.endpoints.gantt_personale import _calcola_gantt_personale

router = APIRouter()

MESI_IT = ['Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
           'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']

STATI_IMPEGNATI = {'attesa_ammin', 'attesa_rs', 'attesa_dir_dip', 'attesa_dg', 'approvata'}


def _mesi_range(data_inizio: date, data_fine: date) -> list[dict]:
    mesi = []
    current = date(data_inizio.year, data_inizio.month, 1)
    fine = date(data_fine.year, data_fine.month, 1)
    while current <= fine:
        mesi.append({
            "key": f"{current.year:04d}-{current.month:02d}",
            "label": f"{MESI_IT[current.month - 1]} {current.year}",
            "anno": current.year,
            "mese": current.month,
        })
        y, m = (current.year, current.month + 1) if current.month < 12 else (current.year + 1, 1)
        current = date(y, m, 1)
    return mesi


def _calcola_andamento(progetto: Progetto, db: Session) -> dict:
    if not progetto.data_inizio or not progetto.data_fine:
        return {"mesi": [], "voci": []}

    mesi = _mesi_range(progetto.data_inizio, progetto.data_fine)
    mesi_keys = {m["key"]: i for i, m in enumerate(mesi)}
    n = len(mesi)

    budget_voci = (
        db.query(BudgetVoce, VoceDiCosto)
        .join(VoceDiCosto, BudgetVoce.voce_id == VoceDiCosto.id)
        .filter(BudgetVoce.progetto_id == progetto.id)
        .filter(BudgetVoce.wp_id.is_(None))  # esclude le sotto-voci WP per evitare doppia somma
        .order_by(VoceDiCosto.codice)
        .all()
    )

    impegni_tutti = db.query(Impegno).filter(Impegno.progetto_id == progetto.id).all()
    spese_tutte = db.query(Spesa).filter(Spesa.progetto_id == progetto.id).all()
    autorizzazioni_tutte = (
        db.query(RichiestaAutorizzazioneSpesa)
        .filter(
            RichiestaAutorizzazioneSpesa.progetto_id == progetto.id,
            RichiestaAutorizzazioneSpesa.stato.in_(STATI_IMPEGNATI),
            RichiestaAutorizzazioneSpesa.budget_voce_id.isnot(None),
        )
        .all()
    )

    aut_per_bv: dict[str, list] = {}
    for a in autorizzazioni_tutte:
        aut_per_bv.setdefault(str(a.budget_voce_id), []).append(a)

    # Dati gantt personale per sub-righe A.1
    try:
        gantt = _calcola_gantt_personale(progetto, db)
        gantt_mesi_idx = {(m["anno"], m["mese"]): i for i, m in enumerate(gantt["mesi"])}
        pian_iniz_mesi = gantt["pianificazione_iniziale"]["mesi"]
        pian_corr_mesi = gantt["pianificazione_corrente"]["mesi"]
    except Exception:
        gantt = None

    result_voci = []
    for bv, vc in budget_voci:
        bv_id = str(bv.id)
        voce_impegni = [i for i in impegni_tutti if str(i.voce_id) == str(bv.voce_id)]
        voce_spese = [s for s in spese_tutte if str(s.voce_id) == str(bv.voce_id)]
        voce_aut = aut_per_bv.get(bv_id, [])

        imp_per_mese = [0.0] * n
        spe_per_mese = [0.0] * n

        for imp in voce_impegni:
            idx = mesi_keys.get(f"{imp.data.year:04d}-{imp.data.month:02d}")
            if idx is not None:
                imp_per_mese[idx] += float(imp.importo)

        for aut in voce_aut:
            dt = aut.created_at.date() if aut.created_at else None
            if dt:
                idx = mesi_keys.get(f"{dt.year:04d}-{dt.month:02d}")
                if idx is not None:
                    imp_per_mese[idx] += float(aut.importo)

        for sp in voce_spese:
            idx = mesi_keys.get(f"{sp.data.year:04d}-{sp.data.month:02d}")
            if idx is not None:
                spe_per_mese[idx] += float(sp.importo)

        # Per A.1 usa i valori PC dal Gantt (per competenza, non per data Spesa)
        if gantt and vc.codice == "A.1":
            for idx, m in enumerate(mesi):
                gi = gantt_mesi_idx.get((m["anno"], m["mese"]))
                if gi is not None and gi < len(pian_corr_mesi):
                    spe_per_mese[idx] = round(float((pian_corr_mesi[gi] or {}).get("costo", 0)), 2)

        # Disponibile calcolato su importo_previsto
        previsto = float(bv.importo_previsto or 0)
        cum_imp = 0.0
        cum_spe = 0.0
        per_mese = []
        for i in range(n):
            cum_imp += imp_per_mese[i]
            cum_spe += spe_per_mese[i]
            per_mese.append({
                "impegnato": round(imp_per_mese[i], 2),
                "speso": round(spe_per_mese[i], 2),
                "disponibile": round(previsto - cum_imp - cum_spe, 2),
            })

        # Solo spese registrate come dettaglio; per A.1 il dettaglio è nei piani_personale
        items = []
        for sp in ([] if vc.codice == "A.1" else voce_spese):
            items.append({
                "tipo": "spesa",
                "descrizione": sp.descrizione or sp.numero_documento or "Spesa",
                "importo": float(sp.importo),
                "col": "speso",
                "mese_key": f"{sp.data.year:04d}-{sp.data.month:02d}",
                "data": sp.data.isoformat(),
            })
        items.sort(key=lambda x: x["data"])

        # Sub-righe pianificazione personale per voce A.1
        piani_personale = None
        if gantt and vc.codice == "A.1":
            def _val_mese(src_mesi, m):
                gi = gantt_mesi_idx.get((m["anno"], m["mese"]))
                return src_mesi[gi] if gi is not None and gi < len(src_mesi) else None

            piani_personale = [
                {
                    "label": "Pianificazione iniziale",
                    "per_mese_spe": [
                        round(float((_val_mese(pian_iniz_mesi, m) or {}).get("costo", 0)), 2)
                        for m in mesi
                    ],
                    "rendicontato": [False] * n,
                },
                {
                    "label": "Pianificazione corrente",
                    "per_mese_spe": [
                        round(float((_val_mese(pian_corr_mesi, m) or {}).get("costo", 0)), 2)
                        for m in mesi
                    ],
                    "rendicontato": [
                        bool((_val_mese(pian_corr_mesi, m) or {}).get("rendicontato", False))
                        for m in mesi
                    ],
                },
            ]

        result_voci.append({
            "id": bv_id,
            "codice": vc.codice,
            "descrizione": vc.descrizione,
            "importo_previsto": previsto,
            "per_mese": per_mese,
            "items": items,
            "piani_personale": piani_personale,
        })

    return {"mesi": mesi, "voci": result_voci}


@router.get("/progetti/{progetto_id}/andamento-mensile")
def andamento_mensile(
    progetto_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    progetto = db.query(Progetto).filter(Progetto.id == progetto_id).first()
    if not progetto:
        raise HTTPException(404, "Progetto non trovato")
    return _calcola_andamento(progetto, db)


@router.get("/progetti/{progetto_id}/andamento-mensile/export/xlsx")
def export_andamento_xlsx(
    progetto_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    progetto = db.query(Progetto).filter(Progetto.id == progetto_id).first()
    if not progetto:
        raise HTTPException(404, "Progetto non trovato")

    data = _calcola_andamento(progetto, db)
    mesi = data["mesi"]
    voci = data["voci"]
    n_mesi = len(mesi)

    wb = Workbook()
    ws = wb.active
    ws.title = "Andamento Mensile"

    thin = Side(style="thin")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    EURO = '#,##0.00" €"'

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def st(cell, font=None, bg=None, align=None, nf=None):
        if font: cell.font = font
        if bg: cell.fill = fill(bg)
        if align: cell.alignment = align
        if nf: cell.number_format = nf
        cell.border = b

    ncols = 2 + 3 * n_mesi
    r = 1

    # Titolo
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    st(ws.cell(r, 1, f"ANDAMENTO MENSILE — {progetto.acronimo or progetto.codice} — {progetto.titolo}"),
       Font(bold=True, color="FFFFFF", size=12), "185FA5", center)
    ws.row_dimensions[r].height = 26
    r += 1

    # Intestazione progetto
    pi_alloc = db.query(Allocazione).filter(
        Allocazione.progetto_id == progetto.id, Allocazione.is_pi == True
    ).first()
    pi_persona = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first() if pi_alloc else None
    ammin_persona = db.query(Persona).filter(Persona.id == progetto.amministrativo_id).first() if progetto.amministrativo_id else None
    dipartimento_nome = progetto.dipartimento.nome if progetto.dipartimento else ""
    pi_nome = f"{pi_persona.cognome} {pi_persona.nome}" if pi_persona else ""
    ammin_nome = f"{ammin_persona.cognome} {ammin_persona.nome}" if ammin_persona else ""
    def fmt_data(d): return d.strftime('%d/%m/%Y') if d else ''

    intestazione_rows = [
        ("Titolo", progetto.titolo or ""),
        ("Acronimo", progetto.acronimo or ""),
        ("Codice", progetto.codice or ""),
        ("CUP", progetto.cup or ""),
        ("Tipo finanziamento", progetto.tipo or ""),
        ("Responsabile scientifico (PI)", pi_nome),
        ("Responsabile amministrativo", ammin_nome),
        ("Dipartimento", dipartimento_nome),
        ("Periodo", f"{fmt_data(progetto.data_inizio)} → {fmt_data(progetto.data_fine)}"),
        ("Fine rendicontazione", fmt_data(progetto.data_fine_rendicontazione)),
        ("Stato", progetto.stato or ""),
        ("Costo totale", float(progetto.costo_totale or 0)),
        ("Importo finanziato", float(progetto.importo_finanziato or 0)),
    ]

    label_font = Font(bold=True, size=9)
    label_bg = "E8EAED"
    value_font = Font(size=9)

    for label, value in intestazione_rows:
        st(ws.cell(r, 1, label), label_font, label_bg, left_al)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
        if isinstance(value, float):
            st(ws.cell(r, 2, value), value_font, "FFFFFF", left_al, EURO)
        else:
            st(ws.cell(r, 2, value), value_font, "FFFFFF", left_al)
        ws.row_dimensions[r].height = 15
        r += 1

    # Riga separatrice vuota
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 6
    r += 1

    # Header riga 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    st(ws.cell(r, 1, "Voce / Item"), Font(bold=True, size=9), "D0D0D0", left_al)
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=2)
    st(ws.cell(r, 2, "Totale\nPrevisto"), Font(bold=True, size=9), "D0D0D0", center)
    for idx, m in enumerate(mesi):
        c0 = 3 + idx * 3
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 2)
        st(ws.cell(r, c0, m["label"]), Font(bold=True, size=9), "D0D0D0", center)
    ws.row_dimensions[r].height = 20
    r += 1

    # Header riga 2: Imp / Spe / Dis
    for idx in range(n_mesi):
        c0 = 3 + idx * 3
        st(ws.cell(r, c0, "Imp."), Font(bold=True, size=8), "E0E0E0", center)
        st(ws.cell(r, c0 + 1, "Spe."), Font(bold=True, size=8), "E0E0E0", center)
        st(ws.cell(r, c0 + 2, "Dis."), Font(bold=True, size=8), "E0E0E0", center)
    ws.row_dimensions[r].height = 16
    r += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    for i in range(n_mesi * 3):
        ws.column_dimensions[get_column_letter(3 + i)].width = 11

    BLU_PI = "DBEAFE"
    BLU_PC = "EDE9FE"

    for voce in voci:
        # Riga voce
        nome = f"{voce['codice']} — {voce['descrizione']}"
        st(ws.cell(r, 1, nome), Font(bold=True, size=10), "F0F0F0", left_al)
        st(ws.cell(r, 2, voce["importo_previsto"]), Font(bold=True, size=9), "F0F0F0", center, EURO)
        for idx, pm in enumerate(voce["per_mese"]):
            c0 = 3 + idx * 3
            st(ws.cell(r, c0, pm["impegnato"] or None), Font(bold=True, size=9), "F0F0F0", center, EURO)
            st(ws.cell(r, c0 + 1, pm["speso"] or None), Font(bold=True, size=9), "F0F0F0", center, EURO)
            dis = pm["disponibile"]
            st(ws.cell(r, c0 + 2, dis),
               Font(bold=True, size=9, color="2E7D32" if dis >= 0 else "C62828"),
               "E8F5E9" if dis >= 0 else "FFEBEE", center, EURO)
        ws.row_dimensions[r].height = 20
        r += 1

        # Sub-righe pianificazione personale (solo A.1)
        if voce.get("piani_personale"):
            for piano in voce["piani_personale"]:
                bg = BLU_PI if "iniziale" in piano["label"] else BLU_PC
                st(ws.cell(r, 1, f"    {piano['label']}"), Font(size=9, italic=True), bg, left_al)
                st(ws.cell(r, 2), bg=bg, align=center)
                for idx, costo in enumerate(piano["per_mese_spe"]):
                    c0 = 3 + idx * 3
                    rend = piano["rendicontato"][idx] if idx < len(piano["rendicontato"]) else False
                    cell_bg = "D1FAE5" if rend else bg
                    st(ws.cell(r, c0), bg=bg, align=center)
                    st(ws.cell(r, c0 + 1, costo or None),
                       Font(size=9, bold=rend), cell_bg, center, EURO)
                    st(ws.cell(r, c0 + 2), bg=bg, align=center)
                ws.row_dimensions[r].height = 16
                r += 1

        # Righe spese dettaglio
        for item in voce["items"]:
            st(ws.cell(r, 1, f"    {item['descrizione']}"), Font(size=9, italic=True), align=left_al)
            st(ws.cell(r, 2), align=center)
            for idx in range(n_mesi):
                c0 = 3 + idx * 3
                m_key = mesi[idx]["key"]
                if item["mese_key"] == m_key:
                    st(ws.cell(r, c0), align=center)
                    st(ws.cell(r, c0 + 1, item["importo"]), Font(size=9), "FFF9C4", center, EURO)
                else:
                    st(ws.cell(r, c0), align=center)
                    st(ws.cell(r, c0 + 1), align=center)
                st(ws.cell(r, c0 + 2), align=center)
            ws.row_dimensions[r].height = 16
            r += 1

    # Riga totali
    NAVY = "1D3557"
    tot_previsto = sum(float(v["importo_previsto"]) for v in voci)
    tot_imp = [sum(float(v["per_mese"][i]["impegnato"]) for v in voci) for i in range(n_mesi)]
    tot_spe = [sum(float(v["per_mese"][i]["speso"]) for v in voci) for i in range(n_mesi)]
    tot_dis = [sum(float(v["per_mese"][i]["disponibile"]) for v in voci) for i in range(n_mesi)]

    st(ws.cell(r, 1, "TOTALE"), Font(bold=True, size=10, color="FFFFFF"), NAVY, left_al)
    st(ws.cell(r, 2, tot_previsto), Font(bold=True, size=9, color="FFFFFF"), NAVY, center, EURO)
    for idx in range(n_mesi):
        c0 = 3 + idx * 3
        st(ws.cell(r, c0, tot_imp[idx] or None), Font(bold=True, size=9, color="FDE68A"), NAVY, center, EURO)
        st(ws.cell(r, c0 + 1, tot_spe[idx] or None), Font(bold=True, size=9, color="FECDD3"), NAVY, center, EURO)
        dis = tot_dis[idx]
        st(ws.cell(r, c0 + 2, dis),
           Font(bold=True, size=9, color="86EFAC" if dis >= 0 else "FCA5A5"), NAVY, center, EURO)
    ws.row_dimensions[r].height = 20
    r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_file = f"andamento_mensile_{progetto.acronimo or progetto.codice}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_file}"},
    )
