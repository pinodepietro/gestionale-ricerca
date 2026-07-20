# backend/app/api/v1/endpoints/timesheet.py
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import and_
from app.core.database import get_db
from app.core.deps import tutti_i_ruoli, richiedi_ruolo
from app.models.persona import Persona
from app.models.timesheet import (
    TimesheetTestata, TimesheetRiga, TimesheetCella,
    TemplateTimesheet, ApprovazioneTimesheet
)
from app.models.personale import CostoOrarioPersona, Allocazione
from app.services.notifiche import crea_notifica, invia_email
from app.models.struttura import WorkPackage
from app.models.progetto import Progetto
from datetime import date, datetime, timezone
import uuid

router = APIRouter()

MESI = {1:"Gennaio",2:"Febbraio",3:"Marzo",4:"Aprile",5:"Maggio",6:"Giugno",
        7:"Luglio",8:"Agosto",9:"Settembre",10:"Ottobre",11:"Novembre",12:"Dicembre"}


def _get_persona_nome(persona_id, db):
    from app.models.persona import Persona
    p = db.query(Persona).filter(Persona.id == persona_id).first()
    return f"{p.cognome} {p.nome}" if p else "—"


def _testata_dict(t: TimesheetTestata, db=None) -> dict:
    ore_progetto = sum(
        float(c.ore)
        for r in t.righe if r.tipo_riga == "progetto"
        for c in r.celle
    )
    ore_totali = sum(float(c.ore) for r in t.righe for c in r.celle)
    return {
        "id": str(t.id),
        "persona_id": str(t.persona_id),
        "progetto_id": str(t.progetto_id),
        "template_id": str(t.template_id),
        "anno": t.anno,
        "mese": t.mese,
        "mese_label": MESI.get(t.mese, ""),
        "sal_id": str(t.sal_id) if t.sal_id else None,
        "stato": t.stato,
        "persona_nome": _get_persona_nome(t.persona_id, db),
        "granularita": t.granularita if hasattr(t, "granularita") else "mensile",
        "inviato_at": t.inviato_at.isoformat() if t.inviato_at else None,
        "approvato_at": t.approvato_at.isoformat() if t.approvato_at else None,
        "ore_totali_progetto": ore_progetto,
        "ore_totali": ore_totali,
        "righe": [_riga_dict(r) for r in t.righe],
    }


def _riga_dict(r: TimesheetRiga) -> dict:
    return {
        "id": str(r.id),
        "testata_id": str(r.testata_id),
        "tipo_riga": r.tipo_riga,
        "wp_id": str(r.wp_id) if r.wp_id else None,
        "task_id": str(r.task_id) if r.task_id else None,
        "progetto_correlato_id": str(r.progetto_correlato_id) if r.progetto_correlato_id else None,
        "descrizione_libera": r.descrizione_libera,
        "ordine": r.ordine,
        "celle": [_cella_dict(c) for c in r.celle],
    }


def _cella_dict(c: TimesheetCella) -> dict:
    return {
        "id": str(c.id),
        "riga_id": str(c.riga_id),
        "giorno": c.giorno,
        "ore": float(c.ore),
        "costo_orario_applicato": float(c.costo_orario_applicato) if c.costo_orario_applicato else None,
        "costo_calcolato": float(c.costo_calcolato) if c.costo_calcolato else None,
    }


def _get_testata_or_404(id: str, db: Session) -> TimesheetTestata:
    t = db.query(TimesheetTestata).filter(TimesheetTestata.id == id).first()
    if not t:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Timesheet non trovato"}})
    return t


# ─── Lista ───────────────────────────────────────────────────────────────────

@router.get("")
def lista_timesheet(
    persona_id: str = Query(None),
    progetto_id: str = Query(None),
    anno: int = Query(None),
    mese: int = Query(None),
    stato: str = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    q = db.query(TimesheetTestata)
    if utente.ruolo == "ricercatore":
        # Ricercatore vede i propri + tutti i timesheet dei progetti in cui è PI
        progetti_pi_sq = db.query(Allocazione.progetto_id).filter(
            Allocazione.persona_id == utente.id,
            Allocazione.is_pi == True,
        ).subquery()
        from sqlalchemy import or_
        q = q.filter(or_(
            TimesheetTestata.persona_id == utente.id,
            TimesheetTestata.progetto_id.in_(progetti_pi_sq),
        ))
    if persona_id:
        q = q.filter(TimesheetTestata.persona_id == persona_id)
    if progetto_id:
        q = q.filter(TimesheetTestata.progetto_id == progetto_id)
    if anno:
        q = q.filter(TimesheetTestata.anno == anno)
    if mese:
        q = q.filter(TimesheetTestata.mese == mese)
    if stato:
        q = q.filter(TimesheetTestata.stato == stato)
    q = q.order_by(TimesheetTestata.anno.desc(), TimesheetTestata.mese.desc())
    total = q.count()
    items = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "data": [_testata_dict(t, db) for t in items],
        "meta": {"total": total, "page": page, "page_size": page_size,
                 "total_pages": max(1, (total + page_size - 1) // page_size)},
    }


# ─── Get singolo ─────────────────────────────────────────────────────────────

@router.get("/{id}")
def get_timesheet(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    return {"data": _testata_dict(t, db)}


# ─── Crea ────────────────────────────────────────────────────────────────────

@router.post("")
def crea_timesheet(
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    progetto_id = body.get("progetto_id")
    anno = body.get("anno")
    mese = body.get("mese")
    granularita = body.get("granularita", "mensile")

    if not all([progetto_id, anno, mese]):
        raise HTTPException(status_code=422, detail={"error": {"code": "CAMPI_MANCANTI",
            "message": "progetto_id, anno e mese sono obbligatori"}})

    # Solo i ricercatori creano timesheet — non amministrativo/management/monitor
    if utente.ruolo in ("amministrativo", "management", "monitor"):
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Solo i ricercatori possono creare timesheet"}})

    # I ricercatori possono creare solo per sé stessi
    persona_id = body.get("persona_id", str(utente.id))
    if utente.ruolo == "ricercatore" and persona_id != str(utente.id):
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Puoi creare timesheet solo per te stesso"}})

    # Verifica duplicato
    esistente = db.query(TimesheetTestata).filter(
        and_(
            TimesheetTestata.persona_id == persona_id,
            TimesheetTestata.progetto_id == progetto_id,
            TimesheetTestata.anno == anno,
            TimesheetTestata.mese == mese,
        )
    ).first()
    if esistente:
        raise HTTPException(status_code=409, detail={"error": {"code": "TIMESHEET_DUPLICATO",
            "message": f"Esiste già un timesheet per questo mese su questo progetto"}})

    # Recupera template dal progetto
    progetto = db.query(Progetto).filter(Progetto.id == progetto_id).first()
    if not progetto:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Progetto non trovato"}})

    template_id = progetto.template_timesheet_id
    if not template_id:
        # Usa il primo template disponibile come fallback
        template = db.query(TemplateTimesheet).first()
        if not template:
            raise HTTPException(status_code=422, detail={"error": {"code": "NESSUN_TEMPLATE",
                "message": "Nessun template timesheet configurato"}})
        template_id = template.id

    t = TimesheetTestata(
        id=uuid.uuid4(),
        persona_id=persona_id,
        progetto_id=progetto_id,
        template_id=template_id,
        anno=anno,
        mese=mese,
        granularita=granularita,
        stato="bozza",
    )
    db.add(t)
    db.flush()

    # Crea righe di default in base al template
    template = db.query(TemplateTimesheet).filter(TemplateTimesheet.id == template_id).first()
    wp_list = db.query(WorkPackage).filter(WorkPackage.progetto_id == progetto_id).all()

    ordine = 0

    if template and template.righe_wp_task and wp_list:
        for wp in wp_list:
            # Determina le celle in base alla granularità
            if granularita == "giornaliero":
                import calendar
                giorni = calendar.monthrange(anno, mese)[1]
                celle = [{"giorno": g, "ore": 0} for g in range(1, giorni + 1)]
            else:
                celle = [{"giorno": 0, "ore": 0}]

            riga = TimesheetRiga(
                id=uuid.uuid4(),
                testata_id=t.id,
                tipo_riga="progetto",
                wp_id=wp.id,
                descrizione_libera=f"{wp.codice} — {wp.titolo}",
                ordine=ordine,
            )
            db.add(riga)
            db.flush()
            for cella_data in celle:
                db.add(TimesheetCella(
                    id=uuid.uuid4(),
                    riga_id=riga.id,
                    giorno=cella_data["giorno"],
                    ore=0,
                ))
            ordine += 1

    import calendar as _cal
    giorni_mese = _cal.monthrange(anno, mese)[1]

    def _crea_celle_riga(riga_id):
        if granularita == "giornaliero":
            for g in range(1, giorni_mese + 1):
                db.add(TimesheetCella(id=uuid.uuid4(), riga_id=riga_id, giorno=g, ore=0))
        else:
            db.add(TimesheetCella(id=uuid.uuid4(), riga_id=riga_id, giorno=0, ore=0))

    if template and template.riga_altri_progetti:
        riga = TimesheetRiga(id=uuid.uuid4(), testata_id=t.id, tipo_riga="altri_progetti",
                             descrizione_libera="Altri progetti finanziati", ordine=ordine)
        db.add(riga)
        db.flush()
        _crea_celle_riga(riga.id)
        ordine += 1

    if template and template.riga_ordinaria:
        riga = TimesheetRiga(id=uuid.uuid4(), testata_id=t.id, tipo_riga="ordinaria",
                             descrizione_libera="Attività ordinaria / non progettuale", ordine=ordine)
        db.add(riga)
        db.flush()
        _crea_celle_riga(riga.id)
        ordine += 1

    if template and template.riga_assenze:
        riga = TimesheetRiga(id=uuid.uuid4(), testata_id=t.id, tipo_riga="assenze",
                             descrizione_libera="Malattia / Ferie / Permessi", ordine=ordine)
        db.add(riga)
        db.flush()
        _crea_celle_riga(riga.id)

    db.commit()
    db.refresh(t)
    return {"data": _testata_dict(t, db)}


# ─── PUT righe (sostituzione completa) ───────────────────────────────────────

@router.put("/{id}/righe")
def aggiorna_righe(
    id: str,
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    if t.stato not in ("bozza", "rifiutato"):
        raise HTTPException(status_code=409, detail={"error": {"code": "TIMESHEET_NON_MODIFICABILE",
            "message": "Solo i timesheet in bozza possono essere modificati"}})

    # Sostituzione completa: elimina tutto e ricrea
    for riga in t.righe:
        db.delete(riga)
    db.flush()

    for ordine, riga_data in enumerate(body.get("righe", [])):
        riga = TimesheetRiga(
            id=uuid.uuid4(),
            testata_id=t.id,
            tipo_riga=riga_data.get("tipo_riga", "progetto"),
            wp_id=riga_data.get("wp_id"),
            task_id=riga_data.get("task_id"),
            progetto_correlato_id=riga_data.get("progetto_correlato_id"),
            descrizione_libera=riga_data.get("descrizione_libera"),
            ordine=ordine,
        )
        db.add(riga)
        db.flush()
        for cella_data in riga_data.get("celle", []):
            db.add(TimesheetCella(
                id=uuid.uuid4(),
                riga_id=riga.id,
                giorno=cella_data.get("giorno", 0),
                ore=cella_data.get("ore", 0),
            ))

    db.commit()
    db.refresh(t)
    ore_totali = sum(float(c.ore) for r in t.righe for c in r.celle)
    return {"data": {"id": str(t.id), "righe_count": len(t.righe), "ore_totali": ore_totali}}


# ─── Transizioni ─────────────────────────────────────────────────────────────

@router.post("/{id}/invia")
def invia_timesheet(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    if t.stato != "bozza":
        raise HTTPException(status_code=409, detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
            "message": "Solo i timesheet in bozza possono essere inviati"}})
    ore_progetto = sum(float(c.ore) for r in t.righe if r.tipo_riga == "progetto" for c in r.celle)
    if ore_progetto <= 0:
        raise HTTPException(status_code=422, detail={"error": {"code": "ORE_MANCANTI",
            "message": "Il timesheet deve avere almeno un'ora di attività progettuale"}})
    t.stato = "inviato"
    t.inviato_at = datetime.now(timezone.utc)

    # Notifica il PI del progetto
    pi_alloc = db.query(Allocazione).filter(
        Allocazione.progetto_id == t.progetto_id,
        Allocazione.is_pi == True,
    ).first()
    if pi_alloc:
        autore = db.query(Persona).filter(Persona.id == t.persona_id).first()
        autore_nome = f"{autore.nome} {autore.cognome}" if autore else "Un membro del team"
        mese_label = MESI.get(t.mese, str(t.mese))
        titolo = f"Nuovo timesheet da approvare"
        messaggio = f"{autore_nome} ha inviato il timesheet di {mese_label} {t.anno}"
        crea_notifica(db, pi_alloc.persona_id, "timesheet_pendente", titolo, messaggio,
                      link=f"/timesheet/{t.id}", riferimento_id=str(t.id))
        pi_persona = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first()
        if pi_persona and pi_persona.email:
            invia_email(pi_persona.email, titolo, messaggio)

    db.commit()
    db.refresh(t)
    return {"data": _testata_dict(t, db)}


@router.post("/{id}/approva")
def approva_timesheet(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    # Solo il PI del progetto specifico può approvare
    e_pi_del_progetto = db.query(Allocazione).filter(
        Allocazione.persona_id == utente.id,
        Allocazione.progetto_id == t.progetto_id,
        Allocazione.is_pi == True,
    ).first()
    if not e_pi_del_progetto:
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Solo il PI del progetto può approvare i timesheet"}})
    if t.stato != "inviato":
        raise HTTPException(status_code=409, detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
            "message": "Solo i timesheet inviati possono essere approvati"}})

    # Registra approvazione del PI
    approvazione = ApprovazioneTimesheet(
        id=uuid.uuid4(),
        testata_id=t.id,
        approvatore_id=utente.id,
        ruolo_firma=utente.ruolo,
        ordine_firma=1,
        esito="approvato",
        data=datetime.now(timezone.utc),
    )
    db.add(approvazione)

    t.stato = "attesa_dg"

    # Notifica il Direttore Generale
    mese_label = MESI.get(t.mese, str(t.mese))
    autore = db.query(Persona).filter(Persona.id == t.persona_id).first()
    autore_nome = f"{autore.nome} {autore.cognome}" if autore else "Un membro del team"
    titolo_dg = f"Timesheet in attesa di approvazione finale — {mese_label} {t.anno}"
    messaggio_dg = f"Il timesheet di {autore_nome} per {mese_label} {t.anno} è stato approvato dal PI e attende la tua approvazione finale."
    from app.models.persona import Persona as PersonaModel
    dg = db.query(PersonaModel).filter(PersonaModel.ruolo == "direttore_generale").first()
    if dg:
        crea_notifica(db, dg.id, "timesheet_attesa_dg", titolo_dg, messaggio_dg,
                      link=f"/timesheet/{t.id}", riferimento_id=str(t.id))

    db.commit()
    db.refresh(t)
    return {"data": _testata_dict(t, db)}


@router.post("/{id}/rifiuta")
def rifiuta_timesheet(
    id: str,
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)

    # Controlla se è PI o DG
    e_pi_del_progetto = db.query(Allocazione).filter(
        Allocazione.persona_id == utente.id,
        Allocazione.progetto_id == t.progetto_id,
        Allocazione.is_pi == True,
    ).first()
    e_dg = utente.ruolo == "direttore_generale" or utente.ruolo == "superadmin"

    if not e_pi_del_progetto and not e_dg:
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Solo il PI del progetto o il Direttore Generale possono rifiutare i timesheet"}})

    # Controlla lo stato in base a chi sta rifiutando
    if e_pi_del_progetto and t.stato != "inviato":
        raise HTTPException(status_code=409, detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
            "message": "Solo i timesheet inviati possono essere rifiutati dal PI"}})
    if e_dg and t.stato != "attesa_dg":
        raise HTTPException(status_code=409, detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
            "message": "Solo i timesheet in attesa di approvazione finale possono essere rifiutati dal DG"}})
    approvazione = ApprovazioneTimesheet(
        id=uuid.uuid4(),
        testata_id=t.id,
        approvatore_id=utente.id,
        ruolo_firma=utente.ruolo,
        ordine_firma=2 if e_dg else 1,
        esito="rifiutato",
        data=datetime.now(timezone.utc),
        note=body.get("note", ""),
    )
    db.add(approvazione)
    t.stato = "rifiutato"

    # Notifica l'autore del timesheet
    mese_label = MESI.get(t.mese, str(t.mese))
    note_testo = body.get("note", "")
    titolo_rif = f"Timesheet rifiutato — {mese_label} {t.anno}"
    messaggio_rif = f"Il tuo timesheet di {mese_label} {t.anno} è stato rifiutato."
    if note_testo:
        messaggio_rif += f" Motivazione: {note_testo}"
    crea_notifica(db, t.persona_id, "timesheet_rifiutato", titolo_rif, messaggio_rif,
                  link=f"/timesheet/{t.id}", urgente=True, riferimento_id=str(t.id))
    autore = db.query(Persona).filter(Persona.id == t.persona_id).first()
    if autore and autore.email:
        invia_email(autore.email, titolo_rif, messaggio_rif)

    db.commit()
    db.refresh(t)
    return {"data": _testata_dict(t, db)}


@router.post("/{id}/approva-finale")
def approva_finale_timesheet(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    # Solo il Direttore Generale può fare l'approvazione finale
    if utente.ruolo != "direttore_generale" and utente.ruolo != "superadmin":
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Solo il Direttore Generale può approvare definitivamente i timesheet"}})
    if t.stato != "attesa_dg":
        raise HTTPException(status_code=409, detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
            "message": "Solo i timesheet in attesa di approvazione finale possono essere approvati"}})

    # Snapshot costo orario su tutte le celle di tipo progetto
    data_approvazione = date.today()
    for riga in t.righe:
        if riga.tipo_riga == "progetto":
            costo = db.query(CostoOrarioPersona).filter(
                and_(
                    CostoOrarioPersona.persona_id == t.persona_id,
                    CostoOrarioPersona.data_inizio <= data_approvazione,
                )
            ).filter(
                (CostoOrarioPersona.data_fine == None) |
                (CostoOrarioPersona.data_fine >= data_approvazione)
            ).first()
            costo_orario = float(costo.costo_orario) if costo else 0.0
            for cella in riga.celle:
                cella.costo_orario_applicato = costo_orario
                cella.costo_calcolato = float(cella.ore) * costo_orario

    # Registra approvazione finale del DG
    approvazione = ApprovazioneTimesheet(
        id=uuid.uuid4(),
        testata_id=t.id,
        approvatore_id=utente.id,
        ruolo_firma=utente.ruolo,
        ordine_firma=2,
        esito="approvato",
        data=datetime.now(timezone.utc),
    )
    db.add(approvazione)
    # ── Aggiorna importo_rendicontato sulla voce Personale ──────────────────
    from app.models.budget import BudgetVoce, VoceDiCosto
    voce_personale = db.query(VoceDiCosto).filter(
        VoceDiCosto.categoria == "personale"
    ).first()
    if voce_personale:
        costo_totale_ts = sum(
            float(c.costo_calcolato or 0)
            for r in t.righe if r.tipo_riga == "progetto"
            for c in r.celle
        )
        bv = db.query(BudgetVoce).filter(
            BudgetVoce.progetto_id == t.progetto_id,
            BudgetVoce.voce_id == voce_personale.id,
        ).first()
        if bv and costo_totale_ts > 0:
            bv.importo_rendicontato = float(bv.importo_rendicontato or 0) + costo_totale_ts
    # ─────────────────────────────────────────────────────────────────────────

    t.stato = "approvato"
    t.approvato_at = datetime.now(timezone.utc)

    # Notifica l'autore del timesheet
    mese_label = MESI.get(t.mese, str(t.mese))
    titolo_esito = f"Timesheet approvato definitivamente — {mese_label} {t.anno}"
    messaggio_esito = f"Il tuo timesheet di {mese_label} {t.anno} è stato approvato definitivamente dal Direttore Generale."
    crea_notifica(db, t.persona_id, "timesheet_approvato", titolo_esito, messaggio_esito,
                  link=f"/timesheet/{t.id}", riferimento_id=str(t.id))
    autore = db.query(Persona).filter(Persona.id == t.persona_id).first()
    if autore and autore.email:
        invia_email(autore.email, titolo_esito, messaggio_esito)

    db.commit()
    db.refresh(t)
    return {"data": _testata_dict(t, db)}


@router.delete("/{id}")
def elimina_timesheet(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    t = _get_testata_or_404(id, db)
    if t.stato not in ("bozza", "rifiutato") and utente.ruolo != "amministrativo":
        raise HTTPException(status_code=409, detail={"error": {"code": "TIMESHEET_NON_ELIMINABILE",
            "message": "Solo i timesheet in bozza o rifiutati possono essere eliminati"}})
    if utente.ruolo == "ricercatore" and str(t.persona_id) != str(utente.id):
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
            "message": "Puoi eliminare solo i tuoi timesheet"}})
    # Elimina manualmente le approvazioni collegate
    from app.models.timesheet import ApprovazioneTimesheet
    db.query(ApprovazioneTimesheet).filter(ApprovazioneTimesheet.testata_id == t.id).delete()
    db.delete(t)
    db.commit()
    return {"data": {"deleted": True}}


@router.get("/{id}/export/xlsx")
def export_timesheet_xlsx(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io, calendar as cal

    t = _get_testata_or_404(id, db)
    from app.models.progetto import Progetto
    from app.models.persona import Persona as PersonaModel

    progetto = db.query(Progetto).filter(Progetto.id == t.progetto_id).first()
    persona = db.query(PersonaModel).filter(PersonaModel.id == t.persona_id).first()
    template = db.query(TemplateTimesheet).filter(TemplateTimesheet.id == t.template_id).first()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Timesheet {t.mese:02d}-{t.anno}"

    blu = "185FA5"
    grigio = "F5F5F5"
    b = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=blu)
    sf = Font(bold=True, size=10)
    sfill = PatternFill("solid", fgColor=grigio)
    center = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="center")

    def st(cell, font=None, fill=None, align=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        cell.border = b

    is_g = t.granularita == "giornaliero"
    giorni = cal.monthrange(t.anno, t.mese)[1] if is_g else 0
    r = 1

    ws.merge_cells(f"A{r}:H{r}")
    st(ws.cell(r, 1, "TIMESHEET MENSILE"), hf, hfill, center)
    ws.row_dimensions[r].height = 28
    r += 1

    for label, val in [
        ("Cognome e Nome", f"{persona.cognome} {persona.nome}" if persona else "—"),
        ("Progetto", f"{progetto.acronimo or progetto.codice} — {progetto.titolo}" if progetto else "—"),
        ("Mese / Anno", f"{MESI.get(t.mese, '')} {t.anno}"),
        ("Stato", t.stato.upper()),
    ]:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 1).border = b
        ws.merge_cells(f"B{r}:H{r}")
        st(ws.cell(r, 2, val), align=left_al)
        r += 1

    r += 1

    if is_g:
        st(ws.cell(r, 1, "Attività"), sf, sfill, center)
        ws.column_dimensions["A"].width = 35
        for g in range(1, giorni + 1):
            c = ws.cell(r, g + 1, str(g))
            st(c, sf, sfill, center)
            ws.column_dimensions[get_column_letter(g + 1)].width = 4
        col_tot = giorni + 2
        st(ws.cell(r, col_tot, "TOT"), Font(bold=True), sfill, center)
        ws.column_dimensions[get_column_letter(col_tot)].width = 6
        r += 1

        for riga in t.righe:
            ws.cell(r, 1, riga.descrizione_libera or riga.tipo_riga).border = b
            tot_r = 0
            for cella in riga.celle:
                if 1 <= cella.giorno <= giorni:
                    ore = float(cella.ore)
                    tot_r += ore
                    c = ws.cell(r, cella.giorno + 1, ore if ore > 0 else "")
                    st(c, align=center)
            st(ws.cell(r, col_tot, tot_r if tot_r > 0 else ""), Font(bold=True), align=center)
            r += 1
    else:
        for col, h in enumerate(["Attività", "Ore totali mese"], 1):
            st(ws.cell(r, col, h), sf, sfill, center)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        r += 1

        tot = 0
        for riga in t.righe:
            ws.cell(r, 1, riga.descrizione_libera or riga.tipo_riga).border = b
            cella = next((c for c in riga.celle if c.giorno == 0), None)
            ore = float(cella.ore) if cella else 0
            if riga.tipo_riga == "progetto": tot += ore
            st(ws.cell(r, 2, ore if ore > 0 else ""), align=center)
            r += 1

        ws.cell(r, 1, "TOTALE ORE PROGETTO").font = Font(bold=True)
        ws.cell(r, 1).border = b
        st(ws.cell(r, 2, tot), Font(bold=True), align=center)
        r += 1

    r += 2

    if template:
        ws.cell(r, 1, "FIRME").font = Font(bold=True)
        r += 1
        firmatari = [f for f in [template.etichetta_firmatario_1,
                                  template.etichetta_firmatario_2,
                                  template.etichetta_firmatario_3] if f]
        col = 1
        for firma in firmatari:
            ws.cell(r, col, firma).font = Font(bold=True, size=9)
            ws.cell(r, col).border = b
            ws.row_dimensions[r + 1].height = 40
            ws.cell(r + 1, col, "").border = b
            col += 3

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import os as _os
    from app.services.storage import progetto_dir, _safe
    _codice = progetto.codice if progetto else "export"
    _cog = _safe(persona.cognome if persona else "export")
    nome = f"TS_{_cog}_{t.mese:02d}{t.anno}.xlsx"
    _output_dir = progetto_dir(_codice, "timesheet")
    _os.makedirs(_output_dir, exist_ok=True)
    _dst = _os.path.join(_output_dir, nome)
    with open(_dst, "wb") as _fh:
        _fh.write(buf.read())
    t.xlsx_path = _dst
    db.commit()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )


@router.get("/{id}/export/template")
def export_timesheet_con_template(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    """Esporta il timesheet usando il template Excel dell'ente finanziatore."""
    from fastapi.responses import StreamingResponse
    from openpyxl import load_workbook
    import io, os
    from app.models.progetto import Progetto
    from app.models.persona import Persona as PersonaModel
    from sqlalchemy import and_

    t = _get_testata_or_404(id, db)
    template = db.query(TemplateTimesheet).filter(TemplateTimesheet.id == t.template_id).first()

    if not template or not template.file_template_path or not os.path.exists(template.file_template_path):
        raise HTTPException(status_code=422, detail={"error": {
            "code": "TEMPLATE_MANCANTE",
            "message": "Nessun file template caricato per questo tipo di timesheet. Usa l'export Excel standard."
        }})

    progetto = db.query(Progetto).filter(Progetto.id == t.progetto_id).first()
    persona = db.query(PersonaModel).filter(PersonaModel.id == t.persona_id).first()

    # Recupera TUTTI i timesheet approvati dell'anno per questa persona+progetto
    tutti_ts = db.query(TimesheetTestata).filter(
        and_(
            TimesheetTestata.persona_id == t.persona_id,
            TimesheetTestata.progetto_id == t.progetto_id,
            TimesheetTestata.anno == t.anno,
            TimesheetTestata.stato == "approvato",
        )
    ).all()

    # Aggrega ore per mese e tipo riga
    ore_per_mese: dict = {}  # {mese: {tipo_riga: ore}}
    for ts in tutti_ts:
        if ts.mese not in ore_per_mese:
            ore_per_mese[ts.mese] = {"progetto": 0, "altri_progetti": 0, "ordinaria": 0, "assenze": 0}
        for riga in ts.righe:
            for cella in riga.celle:
                ore_per_mese[ts.mese][riga.tipo_riga] = \
                    ore_per_mese[ts.mese].get(riga.tipo_riga, 0) + float(cella.ore)

    # Mappa mese -> colonna Excel (gennaio=B=2, febbraio=C=3, ...)
    mese_a_colonna = {i: i + 1 for i in range(1, 13)}  # mese 1=col 2, mese 12=col 13

    # Carica template e popola
    wb = load_workbook(template.file_template_path)
    ws = wb.active

    # Popola dati anagrafici
    if persona:
        ws['A14'] = f"Nominativo: {persona.cognome} {persona.nome}"
        ws['A15'] = f"Qualifica: {persona.ruolo_ente or persona.ruolo}"
        ws['A16'] = f"Contratto applicato: {persona.livello_contratto or '—'}"
    if progetto:
        ws['A7'] = f"Progetto: {progetto.acronimo or progetto.codice} — {progetto.titolo}"
    ws['B19'] = f"ANNO {t.anno}"

    # Popola ore per mese
    for mese, col in mese_a_colonna.items():
        dati_mese = ore_per_mese.get(mese, {})
        ore_prog = dati_mese.get("progetto", 0)
        ore_altri = dati_mese.get("altri_progetti", 0)
        ore_ord = dati_mese.get("ordinaria", 0)

        if ore_prog > 0:
            ws.cell(21, col, ore_prog)   # Attività progetto
        if ore_altri > 0:
            ws.cell(22, col, ore_altri)  # Altri progetti
        if ore_ord > 0:
            ws.cell(23, col, ore_ord)    # Attività ordinaria

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import os as _os
    from app.services.storage import progetto_dir, _safe
    _codice = progetto.codice if progetto else "export"
    _cog = _safe(persona.cognome if persona else "export")
    nome = f"TS_{_cog}_{t.anno}.xlsx"
    _output_dir = progetto_dir(_codice, "timesheet")
    _os.makedirs(_output_dir, exist_ok=True)
    _dst = _os.path.join(_output_dir, nome)
    with open(_dst, "wb") as _fh:
        _fh.write(buf.read())
    t.xlsx_path = _dst
    db.commit()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )
