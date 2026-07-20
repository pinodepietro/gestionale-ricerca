# backend/app/api/v1/endpoints/progetti.py
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from app.core.database import get_db
from app.core.deps import tutti_i_ruoli, solo_amministrativo, solo_superadmin
from app.models.progetto import Progetto
from app.models.partner import Partner, ProgettoPartner, TipoFinanziamento, Finanziamento
from app.models.struttura import WorkPackage, Milestone, Deliverable
from app.models.budget import VoceDiCosto, BudgetVoce, Spesa, Sal, Impegno
from app.models.personale import Allocazione, MonteOreAnnuale
from app.models.timesheet import TimesheetTestata
from app.models.persona import Persona
from datetime import date, datetime
from fastapi import BackgroundTasks
from app.core.config import settings
import math
import io
import urllib.request
from app.api.v1.endpoints.personale import _trigger_sync_progetti
from app.services.notifiche import crea_notifica


def _notifica_sync_missioni():
    try:
        req = urllib.request.Request(
            f"{settings.MISSIONI_URL}/internal/sync-progetti/",
            data=b"",
            method="POST",
            headers={"X-Sync-Key": settings.SYNC_API_KEY},
        )
        urllib.request.urlopen(req, timeout=5)
    except Exception:
        pass  # non blocca mai l'operazione principale

router = APIRouter()
# Router separato per endpoint che non hanno prefisso /progetti
sub_router = APIRouter()

TRANSIZIONI = {
    "bozza":        {"attiva": "attivo"},
    "attivo":       {"chiudi": "chiuso"},
    "chiuso":       {"attiva": "attivo", "rendiconta": "rendicontato"},
    "rendicontato": {},
}


def pagina(items, total, page, page_size):
    return {"data": items, "meta": {"total": total, "page": page, "page_size": page_size,
            "total_pages": math.ceil(total / page_size) if page_size else 1}}


def progetto_dict(p: Progetto) -> dict:
    return {
        "id": str(p.id), "codice": p.codice, "titolo": p.titolo,
        "acronimo": p.acronimo, "descrizione": p.descrizione, "tipo": p.tipo,
        "stato": p.stato,
        "data_inizio": str(p.data_inizio) if p.data_inizio else None,
        "data_fine": str(p.data_fine) if p.data_fine else None,
        "data_fine_rendicontazione": str(p.data_fine_rendicontazione) if p.data_fine_rendicontazione else None,
        "costo_totale": float(p.costo_totale) if p.costo_totale else 0,
        "importo_finanziato": float(p.importo_finanziato) if p.importo_finanziato else 0,
        "quota_cofinanziamento": float(p.costo_totale or 0) - float(p.importo_finanziato or 0),
        "percentuale_finanziamento": round(float(p.importo_finanziato or 0) / float(p.costo_totale or 1) * 100, 2),
        "cup": p.cup, "budget_per_partner": p.budget_per_partner,
        "gestione_per_wp": bool(p.gestione_per_wp),
        "template_timesheet_id": str(p.template_timesheet_id) if p.template_timesheet_id else None,
        "riferimento_bando": p.riferimento_bando,
        "note": p.note,
        "amministrativo_id": str(p.amministrativo_id) if p.amministrativo_id else None,
        "pi_id": str(p.pi_id) if p.pi_id else None,
        "dipartimento_id": str(p.dipartimento_id) if p.dipartimento_id else None,
        "dipartimento_nome": p.dipartimento.nome if hasattr(p, 'dipartimento') and p.dipartimento else None,
    }


# ─── Progetti ────────────────────────────────────────────────────────────────

@router.get("")
def lista_progetti(
    stato: str = Query(None), tipo: str = Query(None),
    search: str = Query(None), includi_bozze: bool = Query(False),
    solo_allocati: bool = Query(False), amministrativo_id: str = Query(None),
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=500),
    db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli),
):
    q = db.query(Progetto)
    if not includi_bozze:
        q = q.filter(Progetto.stato != "bozza")
    if stato:
        q = q.filter(Progetto.stato == stato)
    if tipo:
        q = q.filter(Progetto.tipo == tipo)
    if search:
        q = q.filter(or_(Progetto.codice.ilike(f"%{search}%"),
                         Progetto.titolo.ilike(f"%{search}%"),
                         Progetto.acronimo.ilike(f"%{search}%")))
    if amministrativo_id:
        q = q.filter(Progetto.amministrativo_id == amministrativo_id)
    elif solo_allocati or utente.ruolo not in ("superadmin", "direttore_generale", "monitor", "management"):
        proj_ids = db.query(Allocazione.progetto_id).filter(Allocazione.persona_id == utente.id).subquery()
        if utente.ruolo == "amministrativo":
            q = q.filter(or_(Progetto.id.in_(proj_ids), Progetto.amministrativo_id == utente.id))
        else:
            q = q.filter(Progetto.id.in_(proj_ids))
    total = q.count()
    items = q.order_by(Progetto.codice).offset((page - 1) * page_size).limit(page_size).all()
    return pagina([progetto_dict(p) for p in items], total, page, page_size)


@router.get("/bozze")
def lista_bozze(
    db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli),
):
    q = db.query(Progetto).filter(Progetto.stato == "bozza")
    # Amministrativo vede solo i progetti assegnati a lui
    if utente.ruolo == "amministrativo":
        q = q.filter(Progetto.amministrativo_id == utente.id)
    items = q.order_by(Progetto.codice).all()
    return {"data": [progetto_dict(p) for p in items]}



# ─── Dashboard / Cruscotto globale ───────────────────────────────────────────

@router.get("/cruscotto")
def cruscotto_globale(
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.budget import Spesa, BudgetVoce
    from app.models.personale import Allocazione
    from app.models.budget import Sal
    from app.models.timesheet import TimesheetTestata
    from datetime import date, timedelta
    from sqlalchemy import func

    # Progetti a cui l'utente partecipa (tramite allocazioni) o tutti se admin/management/monitor
    if utente.ruolo in ("management", "monitor", "superadmin", "direttore_generale"):
        progetti_ids = [str(p.id) for p in db.query(Progetto).filter(Progetto.stato == "attivo").all()]
    elif utente.ruolo == "amministrativo":
        progetti_ids = [str(p.id) for p in db.query(Progetto).filter(
            Progetto.stato == "attivo",
            Progetto.amministrativo_id == utente.id
        ).all()]
    else:
        allocazioni = db.query(Allocazione).filter(Allocazione.persona_id == utente.id).all()
        progetti_ids = list({str(a.progetto_id) for a in allocazioni})

    if not progetti_ids:
        return {"data": {
            "progetti_attivi": 0,
            "budget_previsto": 0,
            "budget_rendicontato": 0,
            "percentuale_budget": 0,
            "spese_totali": 0,
            "timesheet_pendenti": 0,
            "sal_in_scadenza": 0,
            "progetti": [],
        }}

    progetti = db.query(Progetto).filter(
        Progetto.id.in_(progetti_ids),
        Progetto.stato == "attivo",
    ).all()

    # KPI globali
    costo_totale_portfolio = sum(float(p.costo_totale or 0) for p in progetti)
    importo_finanziato_portfolio = sum(float(p.importo_finanziato or 0) for p in progetti)

    # pianificato = somma voci di budget
    budget_pianificato = db.query(func.sum(BudgetVoce.importo_previsto)).filter(
        BudgetVoce.progetto_id.in_(progetti_ids)
    ).scalar() or 0

    budget_rendicontato = db.query(func.sum(BudgetVoce.importo_rendicontato)).filter(
        BudgetVoce.progetto_id.in_(progetti_ids)
    ).scalar() or 0

    spese_totali = db.query(func.sum(Spesa.importo)).filter(
        Spesa.progetto_id.in_(progetti_ids),
        Spesa.stato == "registrata",
    ).scalar() or 0

    # Spese per progetto (escluso personale — quello viene dai timesheet)
    from app.models.budget import VoceDiCosto as VDC
    voci_personale_ids = [
        str(v.id) for v in db.query(VDC).filter(VDC.categoria.in_(["personale", "overhead"])).all()
    ]

    timesheet_pendenti = db.query(TimesheetTestata).filter(
        TimesheetTestata.progetto_id.in_(progetti_ids),
        TimesheetTestata.stato == "inviato",
    ).count()

    oggi = date.today()
    tra_30_giorni = oggi + timedelta(days=30)
    sal_in_scadenza = db.query(Sal).filter(
        Sal.progetto_id.in_(progetti_ids),
        Sal.stato.in_(["aperto", "chiuso"]),
        Sal.data_scadenza_rendiconto != None,
        Sal.data_scadenza_rendiconto <= tra_30_giorni,
        Sal.data_scadenza_rendiconto >= oggi,
    ).count()

    # Dettaglio per progetto
    progetti_detail = []
    for p in progetti:
        bv_rows = db.query(BudgetVoce).filter(BudgetVoce.progetto_id == p.id).all()
        prev = sum(float(b.importo_previsto or 0) for b in bv_rows)
        rend = sum(float(b.importo_rendicontato or 0) for b in bv_rows)

        # Percentuale tempo trascorso
        if p.data_inizio and p.data_fine:
            durata_totale = (p.data_fine - p.data_inizio).days
            trascorso = (oggi - p.data_inizio).days
            pct_tempo = min(100, max(0, round(trascorso / durata_totale * 100, 1))) if durata_totale > 0 else 0
        else:
            pct_tempo = 0

        pi_alloc = db.query(Allocazione).filter(
            Allocazione.progetto_id == p.id,
            Allocazione.is_pi == True,
        ).first()
        pi_nome = None
        if pi_alloc:
            pi_p = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first()
            pi_nome = f"{pi_p.nome} {pi_p.cognome}" if pi_p else None

        progetti_detail.append({
            "id": str(p.id),
            "codice": p.codice,
            "acronimo": p.acronimo or p.codice,
            "titolo": p.titolo,
            "tipo": p.tipo,
            "data_inizio": str(p.data_inizio) if p.data_inizio else None,
            "data_fine": str(p.data_fine) if p.data_fine else None,
            "pianificato": prev,
            "rendicontato": rend,
            "pct_rendicontato": round(rend / prev * 100, 1) if prev > 0 else 0,
            "pct_speso": round(
                float(db.query(func.sum(Spesa.importo)).filter(
                    Spesa.progetto_id == p.id, Spesa.stato == "registrata"
                ).scalar() or 0) / float(p.importo_finanziato) * 100, 1
            ) if p.importo_finanziato else 0,
            "percentuale_tempo": pct_tempo,
            "importo_finanziato": float(p.importo_finanziato) if p.importo_finanziato else 0,
            "costo_totale": float(p.costo_totale) if p.costo_totale else 0,
            "budget_allocato_voci": float(
                db.query(func.sum(BudgetVoce.importo_previsto)).filter(
                    BudgetVoce.progetto_id == p.id
                ).scalar() or 0
            ),
            "spese_documentate": float(
                db.query(func.sum(Spesa.importo)).filter(
                    Spesa.progetto_id == p.id,
                    Spesa.stato == "registrata",
                    Spesa.voce_id.notin_(voci_personale_ids) if voci_personale_ids else True,
                ).scalar() or 0
            ),
            "budget_spese_ammissibili": float(
                db.query(func.sum(BudgetVoce.importo_previsto)).join(
                    VoceDiCosto, BudgetVoce.voce_id == VoceDiCosto.id
                ).filter(
                    BudgetVoce.progetto_id == p.id,
                    VoceDiCosto.categoria.notin_(["personale", "overhead"]),
                ).scalar() or 0
            ),
            "pi_nome": pi_nome,
        })

    return {"data": {
        "progetti_attivi": len(progetti),
        "costo_totale_portfolio": costo_totale_portfolio,
        "importo_finanziato_portfolio": importo_finanziato_portfolio,
        "budget_pianificato": float(budget_pianificato),
        "budget_rendicontato": float(budget_rendicontato),
        "pct_rendicontato": round(float(budget_rendicontato) / float(budget_pianificato) * 100, 1) if budget_pianificato > 0 else 0,
        "spese_totali": float(spese_totali),
        "pct_speso": round(float(spese_totali) / importo_finanziato_portfolio * 100, 1) if importo_finanziato_portfolio > 0 else 0,
        "timesheet_pendenti": timesheet_pendenti,
        "sal_in_scadenza": sal_in_scadenza,
        "progetti": progetti_detail,
    }}


@router.get("/cruscotto-dg")
def cruscotto_direttore_generale(
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    """Cruscotto del Direttore Generale con approvazioni da fare"""
    from app.models.timesheet import TimesheetTestata
    from app.models.missione import Missione, RimborsoMissione
    from app.models.rimborso_spesa import RichiestaRimborsoSpesa
    from app.models.autorizzazione_spesa import RichiestaAutorizzazioneSpesa

    if utente.ruolo != "direttore_generale" and utente.ruolo != "superadmin":
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN"}})

    # Conteggi approvazioni in attesa + ID primo elemento
    ts_query = db.query(TimesheetTestata).filter(TimesheetTestata.stato == "attesa_dg")
    ts_primo = ts_query.first()
    timesheet_da_approvare = ts_query.count()
    timesheet_primo_id = str(ts_primo.id) if ts_primo else None

    m_query = db.query(Missione).filter(Missione.stato == "attesa_dg")
    m_primo = m_query.first()
    missioni_da_approvare = m_query.count()
    missioni_primo_id = str(m_primo.id) if m_primo else None

    rm_query = db.query(RimborsoMissione).filter(RimborsoMissione.stato == "attesa_dg")
    rm_primo = rm_query.first()
    rimborsi_missione_da_approvare = rm_query.count()
    rimborsi_missione_primo_id = str(rm_primo.id) if rm_primo else None

    rs_query = db.query(RichiestaRimborsoSpesa).filter(RichiestaRimborsoSpesa.stato == "attesa_dg")
    rs_primo = rs_query.first()
    rimborsi_spesa_da_approvare = rs_query.count()
    rimborsi_spesa_primo_id = str(rs_primo.id) if rs_primo else None

    as_query = db.query(RichiestaAutorizzazioneSpesa).filter(RichiestaAutorizzazioneSpesa.stato == "attesa_dg")
    as_primo = as_query.first()
    autorizzazioni_spesa_da_approvare = as_query.count()
    autorizzazioni_spesa_primo_id = str(as_primo.id) if as_primo else None

    totale_approvazioni = (
        timesheet_da_approvare +
        missioni_da_approvare +
        rimborsi_missione_da_approvare +
        rimborsi_spesa_da_approvare +
        autorizzazioni_spesa_da_approvare
    )

    return {"data": {
        "timesheet": timesheet_da_approvare,
        "timesheet_primo_id": timesheet_primo_id,
        "missioni": missioni_da_approvare,
        "missioni_primo_id": missioni_primo_id,
        "rimborsi_missione": rimborsi_missione_da_approvare,
        "rimborsi_missione_primo_id": rimborsi_missione_primo_id,
        "rimborsi_spesa": rimborsi_spesa_da_approvare,
        "rimborsi_spesa_primo_id": rimborsi_spesa_primo_id,
        "autorizzazioni_spesa": autorizzazioni_spesa_da_approvare,
        "autorizzazioni_spesa_primo_id": autorizzazioni_spesa_primo_id,
        "totale": totale_approvazioni,
    }}


@router.get("/{id}")
def get_progetto(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    p = _get_or_404(id, db)
    d = progetto_dict(p)
    pi_alloc = db.query(Allocazione).filter(
        Allocazione.progetto_id == p.id,
        Allocazione.is_pi == True,
    ).first()
    if pi_alloc:
        pi = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first()
        d["pi_nome"] = f"{pi.nome} {pi.cognome}" if pi else None
    else:
        d["pi_nome"] = None
    if p.amministrativo_id:
        amm = db.query(Persona).filter(Persona.id == p.amministrativo_id).first()
        d["amministrativo_nome"] = f"{amm.nome} {amm.cognome}" if amm else None
    else:
        d["amministrativo_nome"] = None
    return {"data": d}


@router.post("")
def crea_progetto(body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_superadmin)):
    campi = {k: v for k, v in body.items() if hasattr(Progetto, k)}
    p = Progetto(**campi)
    p.stato = "bozza"
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"data": progetto_dict(p)}


@router.patch("/{id}")
def aggiorna_progetto(id: str, body: dict, background_tasks: BackgroundTasks, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    p = _get_or_404(id, db)
    for k, v in body.items():
        if hasattr(Progetto, k) and k not in ("id", "stato"):
            setattr(p, k, v)
    db.commit()
    db.refresh(p)
    if p.stato != "bozza":
        background_tasks.add_task(_notifica_sync_missioni)
    return {"data": progetto_dict(p)}


@router.post("/{id}/attiva")
def attiva(id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    p = _get_or_404(id, db)
    errori = []

    ha_wp = db.query(WorkPackage).filter(WorkPackage.progetto_id == id).first() is not None
    if not ha_wp:
        errori.append("Il progetto deve avere almeno un Work Package")

    ha_allocazioni = db.query(Allocazione).filter(Allocazione.progetto_id == id).first() is not None
    if not ha_allocazioni:
        errori.append("Il progetto deve avere almeno un membro del personale allocato")

    budget_totale = db.query(BudgetVoce).filter(BudgetVoce.progetto_id == id).first()
    if not budget_totale:
        errori.append("Il progetto deve avere almeno una voce di budget definita")

    if errori:
        raise HTTPException(
            status_code=422,
            detail={"error": {
                "code": "PRE_ATTIVAZIONE_FALLITA",
                "message": "Il progetto non può essere attivato",
                "dettagli": errori,
            }},
        )

    result = _transizione(id, "attiva", db)
    background_tasks.add_task(_notifica_sync_missioni)
    return result


@router.post("/{id}/chiudi")
def chiudi(id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    return _transizione(id, "chiudi", db)


# ─── Partner del progetto ─────────────────────────────────────────────────────

@router.get("/{id}/partner")
def lista_partner_progetto(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    _get_or_404(id, db)
    pp = db.query(ProgettoPartner).filter(ProgettoPartner.progetto_id == id).all()
    return {"data": [{"id": str(x.id), "progetto_id": str(x.progetto_id),
                      "partner_id": str(x.partner_id), "ruolo": x.ruolo,
                      "budget_assegnato": float(x.budget_assegnato) if x.budget_assegnato else None,
                      "partner": {"id": str(x.partner.id), "nome": x.partner.nome} if x.partner else None}
                     for x in pp]}


@router.post("/{id}/partner")
def aggiungi_partner(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    _get_or_404(id, db)
    pp = ProgettoPartner(progetto_id=id, partner_id=body["partner_id"],
                         ruolo=body.get("ruolo", "partner"),
                         budget_assegnato=body.get("budget_assegnato"))
    db.add(pp)
    db.commit()
    db.refresh(pp)
    return {"data": {"id": str(pp.id), "partner_id": str(pp.partner_id), "ruolo": pp.ruolo}}


@router.delete("/{id}/partner/{pp_id}")
def rimuovi_partner(id: str, pp_id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    pp = db.query(ProgettoPartner).filter(ProgettoPartner.id == pp_id, ProgettoPartner.progetto_id == id).first()
    if not pp:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Partner non trovato"}})
    db.delete(pp)
    db.commit()
    return {"data": {"deleted": True}}


# ─── Budget voci ──────────────────────────────────────────────────────────────

@router.get("/{id}/budget")
def lista_budget(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    from sqlalchemy import func as sqlfunc
    _get_or_404(id, db)
    voci = db.query(BudgetVoce).filter(BudgetVoce.progetto_id == id).all()
    # speso per voce (da tabella spesa)
    spese_per_voce = dict(
        db.query(Spesa.voce_id, sqlfunc.sum(Spesa.importo))
        .filter(Spesa.progetto_id == id, Spesa.stato == "registrata")
        .group_by(Spesa.voce_id)
        .all()
    )
    return {"data": [{"id": str(v.id), "progetto_id": str(v.progetto_id),
                      "voce_id": str(v.voce_id),
                      "wp_id": str(v.wp_id) if v.wp_id else None,
                      "voce": {"codice": v.voce.codice, "descrizione": v.voce.descrizione, "categoria": v.voce.categoria} if v.voce else None,
                      "importo_previsto": float(v.importo_previsto),
                      "importo_erogato": float(v.importo_erogato or 0),
                      "importo_rendicontato": float(v.importo_rendicontato),
                      "importo_impegnato": float(v.importo_impegnato),
                      "importo_speso": float(spese_per_voce.get(v.voce_id, 0)),
                      "importo_disponibile": float(v.importo_erogato or 0) - float(v.importo_impegnato) - float(spese_per_voce.get(v.voce_id, 0)),
                      "importo_residuo": float(v.importo_previsto) - float(v.importo_rendicontato),
                      "percentuale_utilizzata": round(float(v.importo_rendicontato) / float(v.importo_previsto) * 100, 2)
                                                if float(v.importo_previsto) > 0 else 0}
                     for v in voci]}


@router.post("/{id}/budget")
def salva_budget(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from sqlalchemy import func as sqlfunc
    p = _get_or_404(id, db)
    voci_input = body.get("voci", [])

    # Constraint 1: importo >= speso + impegnato per ogni voce esistente
    existing_map = {str(bv.voce_id): bv for bv in db.query(BudgetVoce).filter(BudgetVoce.progetto_id == id).all()}
    spese_per_voce = dict(
        db.query(Spesa.voce_id, sqlfunc.sum(Spesa.importo))
        .filter(Spesa.progetto_id == id, Spesa.stato == "registrata")
        .group_by(Spesa.voce_id).all()
    )
    for voce_data in voci_input:
        vid = voce_data["voce_id"]
        if vid in existing_map:
            bv = existing_map[vid]
            min_importo = float(bv.importo_impegnato or 0) + float(spese_per_voce.get(bv.voce_id, 0) or 0)
            nuovo_importo = float(voce_data.get("importo_previsto", 0))
            if nuovo_importo < min_importo - 0.01:
                raise HTTPException(
                    status_code=422,
                    detail={"error": {
                        "code": "IMPORTO_SOTTO_MINIMO",
                        "message": f"Importo previsto ({nuovo_importo:,.2f}€) inferiore a speso+impegnato ({min_importo:,.2f}€) per la voce",
                    }},
                )

    # Constraint 2: totale voci deve essere uguale al costo totale del progetto
    totale_voci = sum(float(v.get("importo_previsto", 0)) for v in voci_input)
    if abs(totale_voci - float(p.costo_totale)) > 0.01:
        raise HTTPException(
            status_code=422,
            detail={"error": {
                "code": "BUDGET_NON_BILANCIA",
                "message": f"Il totale delle voci ({totale_voci:,.2f}€) deve essere uguale al costo totale del progetto ({float(p.costo_totale):,.2f}€)",
            }},
        )
    voci_input = body.get("voci", [])
    # Mantieni solo i record che corrispondono all'elenco inviato (match per id se presente)
    ids_inviati = [v["id"] for v in voci_input if v.get("id")]
    if ids_inviati:
        db.query(BudgetVoce).filter(
            BudgetVoce.progetto_id == id,
            BudgetVoce.id.notin_(ids_inviati)
        ).delete(synchronize_session=False)
    else:
        # Fallback compatibilità: rimuovi per voce_id (progetti senza WP)
        voci_ids = [v["voce_id"] for v in voci_input]
        db.query(BudgetVoce).filter(
            BudgetVoce.progetto_id == id,
            BudgetVoce.voce_id.notin_(voci_ids)
        ).delete(synchronize_session=False)
    # Aggiorna o crea le voci presenti
    for voce_data in voci_input:
        if voce_data.get("id"):
            existing = db.query(BudgetVoce).filter(BudgetVoce.id == voce_data["id"]).first()
            if existing:
                existing.importo_previsto = voce_data["importo_previsto"]
                existing.wp_id = voce_data.get("wp_id")
                continue
        # Cerca per (progetto, voce, wp)
        q = db.query(BudgetVoce).filter(
            BudgetVoce.progetto_id == id,
            BudgetVoce.voce_id == voce_data["voce_id"],
        )
        if voce_data.get("wp_id"):
            q = q.filter(BudgetVoce.wp_id == voce_data["wp_id"])
        else:
            q = q.filter(BudgetVoce.wp_id.is_(None))
        existing = q.first()
        if existing:
            existing.importo_previsto = voce_data["importo_previsto"]
        else:
            bv = BudgetVoce(progetto_id=id, voce_id=voce_data["voce_id"],
                            wp_id=voce_data.get("wp_id"),
                            importo_previsto=voce_data["importo_previsto"],
                            importo_rendicontato=0)
            db.add(bv)
    db.commit()
    return {"data": {"saved": True}}


# ─── Work Package ─────────────────────────────────────────────────────────────

@router.get("/{id}/wp")
def lista_wp(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    _get_or_404(id, db)
    wps = db.query(WorkPackage).filter(WorkPackage.progetto_id == id).order_by(WorkPackage.codice).all()
    return {"data": [{"id": str(w.id), "progetto_id": str(w.progetto_id),
                      "codice": w.codice, "titolo": w.titolo, "descrizione": w.descrizione,
                      "data_inizio": str(w.data_inizio), "data_fine": str(w.data_fine),
                      "stato": w.stato, "responsabile_id": str(w.responsabile_id) if w.responsabile_id else None}
                     for w in wps]}


@router.post("/{id}/wp")
def crea_wp(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    p = _get_or_404(id, db)
    data_inizio = date.fromisoformat(body["data_inizio"])
    data_fine = date.fromisoformat(body["data_fine"])
    _valida_date_nel_progetto(p, data_inizio, data_fine, "Work Package")
    wp = WorkPackage(progetto_id=id, codice=body["codice"], titolo=body["titolo"],
                     descrizione=body.get("descrizione"),
                     data_inizio=data_inizio,
                     data_fine=data_fine,
                     stato=body.get("stato", "pianificato"),
                     responsabile_id=body.get("responsabile_id"))
    db.add(wp)
    db.commit()
    db.refresh(wp)
    return {"data": {"id": str(wp.id), "codice": wp.codice, "titolo": wp.titolo}}


@sub_router.patch("/wp/{wp_id}")
def aggiorna_wp(wp_id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    wp = db.query(WorkPackage).filter(WorkPackage.id == wp_id).first()
    if not wp:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "WP non trovato"}})
    
    # Validazione date se vengono modificate
    data_inizio = date.fromisoformat(body["data_inizio"]) if "data_inizio" in body else wp.data_inizio
    data_fine = date.fromisoformat(body["data_fine"]) if "data_fine" in body else wp.data_fine
    progetto = db.query(Progetto).filter(Progetto.id == wp.progetto_id).first()
    if progetto:
        _valida_date_nel_progetto(progetto, data_inizio, data_fine, "Work Package")
    
    for k, v in body.items():
        if hasattr(WorkPackage, k) and k != "id":
            setattr(wp, k, v)
    db.commit()
    db.refresh(wp)
    return {"data": {"id": str(wp.id), "codice": wp.codice}}


@sub_router.delete("/wp/{wp_id}")
def elimina_wp(wp_id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    wp = db.query(WorkPackage).filter(WorkPackage.id == wp_id).first()
    if not wp:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "WP non trovato"}})
    db.delete(wp)
    db.commit()
    return {"data": {"deleted": True}}


# ─── Allocazioni ─────────────────────────────────────────────────────────────

@router.get("/{id}/allocazioni")
def lista_allocazioni(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    from app.models.persona import Persona as PersonaModel
    from app.models.personale import CostoOrarioPersona
    from sqlalchemy import or_
    _get_or_404(id, db)
    alloc = db.query(Allocazione).join(PersonaModel, Allocazione.persona_id == PersonaModel.id).filter(Allocazione.progetto_id == id).order_by(PersonaModel.cognome, PersonaModel.nome).all()

    def _costo_orario(persona_id, data_rif):
        rate = db.query(CostoOrarioPersona).filter(
            CostoOrarioPersona.persona_id == persona_id,
            CostoOrarioPersona.data_inizio <= data_rif,
            or_(CostoOrarioPersona.data_fine.is_(None), CostoOrarioPersona.data_fine >= data_rif),
        ).order_by(CostoOrarioPersona.data_inizio.desc()).first()
        if rate:
            return float(rate.costo_orario)
        # fallback: tariffa più vicina disponibile (allocazione antecedente all'inserimento del costo)
        fallback = db.query(CostoOrarioPersona).filter(
            CostoOrarioPersona.persona_id == persona_id,
        ).order_by(CostoOrarioPersona.data_inizio.asc()).first()
        return float(fallback.costo_orario) if fallback else 0.0

    return {"data": [{"id": str(a.id), "persona_id": str(a.persona_id),
                      "progetto_id": str(a.progetto_id),
                      "wp_id": str(a.wp_id) if a.wp_id else None,
                      "ore_assegnate": float(a.ore_assegnate),
                      "data_inizio": str(a.data_inizio), "data_fine": str(a.data_fine),
                      "note": a.note,
                      "is_pi": bool(a.is_pi),
                      "is_ammin": bool(a.is_ammin),
                      "costo_orario": _costo_orario(a.persona_id, a.data_inizio),
                      "persona": {"nome": a.persona.nome, "cognome": a.persona.cognome} if a.persona else None}
                     for a in alloc]}


@router.post("/{id}/allocazioni")
def crea_allocazione(id: str, body: dict, background_tasks: BackgroundTasks, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    p = _get_or_404(id, db)
    data_inizio = date.fromisoformat(body["data_inizio"])
    data_fine = date.fromisoformat(body["data_fine"])
    _valida_date_nel_progetto(p, data_inizio, data_fine, "Allocazione personale")
    anno = data_inizio.year

    # Per le sotto-allocazioni WP le ore sono già conteggiate nell'allocazione di progetto: skip monte ore
    if not body.get("wp_id"):
        monte = db.query(MonteOreAnnuale).filter(
            MonteOreAnnuale.persona_id == body["persona_id"],
            MonteOreAnnuale.anno == anno,
        ).first()
        if monte:
            ore_residue = float(monte.ore_disponibili) - float(monte.ore_allocate)
            if float(body["ore_assegnate"]) > ore_residue:
                raise HTTPException(
                    status_code=422,
                    detail={"error": {"code": "MONTE_ORE_INSUFFICIENTE",
                                      "message": f"Ore residue disponibili: {ore_residue}h",
                                      "detail": {"ore_residue": ore_residue}}},
                )
            monte.ore_allocate = float(monte.ore_allocate) + float(body["ore_assegnate"])

    # Controlla duplicati: unicità per (progetto, persona, wp) o (progetto, persona) se wp=null
    q_dup = db.query(Allocazione).filter(
        Allocazione.progetto_id == id,
        Allocazione.persona_id == body["persona_id"],
    )
    if body.get("wp_id"):
        q_dup = q_dup.filter(Allocazione.wp_id == body["wp_id"])
    else:
        q_dup = q_dup.filter(Allocazione.wp_id.is_(None))
    esistente = q_dup.first()
    if esistente:
        if body.get("wp_id"):
            raise HTTPException(status_code=409, detail={"error": {
                "code": "PERSONA_GIA_ALLOCATA_SU_WP",
                "message": "Questa persona è già allocata su questo Work Package."
            }})
        else:
            raise HTTPException(status_code=409, detail={"error": {
                "code": "PERSONA_GIA_ALLOCATA",
                "message": "Questa persona è già allocata su questo progetto. Modifica l'allocazione esistente."
            }})

    # Se is_pi=True, rimuovi il ruolo PI da eventuali allocazioni precedenti
    if body.get("is_pi"):
        db.query(Allocazione).filter(
            Allocazione.progetto_id == id,
            Allocazione.is_pi == True
        ).update({"is_pi": False})

    # Se is_ammin=True, rimuovi il ruolo AMMIN da eventuali allocazioni precedenti
    if body.get("is_ammin"):
        db.query(Allocazione).filter(
            Allocazione.progetto_id == id,
            Allocazione.is_ammin == True
        ).update({"is_ammin": False})

    a = Allocazione(progetto_id=id, persona_id=body["persona_id"],
                    wp_id=body.get("wp_id"),
                    ore_assegnate=body["ore_assegnate"],
                    data_inizio=date.fromisoformat(body["data_inizio"]),
                    data_fine=date.fromisoformat(body["data_fine"]),
                    note=body.get("note"),
                    is_pi=bool(body.get("is_pi", False)),
                    is_ammin=bool(body.get("is_ammin", False)))
    db.add(a)
    db.commit()
    db.refresh(a)
    if a.is_pi or a.is_ammin:
        background_tasks.add_task(_trigger_sync_progetti)
    return {"data": {"id": str(a.id), "ore_assegnate": float(a.ore_assegnate)}}


@sub_router.patch("/progetti/{id}/allocazioni/{alloc_id}")
def aggiorna_allocazione(id: str, alloc_id: str, body: dict, background_tasks: BackgroundTasks, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    a = db.query(Allocazione).filter(Allocazione.id == alloc_id, Allocazione.progetto_id == id).first()
    if not a:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Allocazione non trovata"}})

    # Validazione date se modificate
    p = _get_or_404(id, db)
    data_inizio = date.fromisoformat(body["data_inizio"]) if "data_inizio" in body else a.data_inizio
    data_fine = date.fromisoformat(body["data_fine"]) if "data_fine" in body else a.data_fine
    _valida_date_nel_progetto(p, data_inizio, data_fine, "Allocazione personale")

    # Aggiorna monte ore se cambiano le ore assegnate
    if "ore_assegnate" in body and float(body["ore_assegnate"]) != float(a.ore_assegnate):
        anno = data_inizio.year
        monte = db.query(MonteOreAnnuale).filter(
            MonteOreAnnuale.persona_id == a.persona_id,
            MonteOreAnnuale.anno == anno,
        ).first()
        if monte:
            differenza = float(body["ore_assegnate"]) - float(a.ore_assegnate)
            if differenza > 0:
                ore_residue = float(monte.ore_disponibili) - float(monte.ore_allocate)
                if differenza > ore_residue:
                    raise HTTPException(
                        status_code=422,
                        detail={"error": {"code": "MONTE_ORE_INSUFFICIENTE",
                                          "message": f"Ore residue disponibili: {ore_residue}h"}},
                    )
            monte.ore_allocate = max(0, float(monte.ore_allocate) + differenza)

    # Se is_pi=True, rimuovi PI da tutte le altre allocazioni del progetto
    if body.get("is_pi"):
        db.query(Allocazione).filter(
            Allocazione.progetto_id == id,
            Allocazione.is_pi == True,
            Allocazione.id != alloc_id,
        ).update({"is_pi": False})

    # Se is_ammin=True, rimuovi AMMIN da tutte le altre allocazioni del progetto
    if body.get("is_ammin"):
        db.query(Allocazione).filter(
            Allocazione.progetto_id == id,
            Allocazione.is_ammin == True,
            Allocazione.id != alloc_id,
        ).update({"is_ammin": False})

    for k, v in body.items():
        if hasattr(Allocazione, k) and k not in ("id", "persona_id", "progetto_id"):
            setattr(a, k, v)
    db.commit()
    db.refresh(a)
    if "is_pi" in body or "is_ammin" in body:
        background_tasks.add_task(_trigger_sync_progetti)
    return {"data": {"id": str(a.id), "ore_assegnate": float(a.ore_assegnate)}}


@sub_router.delete("/progetti/{id}/allocazioni/{alloc_id}")
def elimina_allocazione(id: str, alloc_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    a = db.query(Allocazione).filter(Allocazione.id == alloc_id, Allocazione.progetto_id == id).first()
    if not a:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Allocazione non trovata"}})
    # Aggiorna monte ore
    anno = a.data_inizio.year
    monte = db.query(MonteOreAnnuale).filter(
        MonteOreAnnuale.persona_id == a.persona_id,
        MonteOreAnnuale.anno == anno,
    ).first()
    if monte:
        monte.ore_allocate = max(0, float(monte.ore_allocate) - float(a.ore_assegnate))
    era_pi = a.is_pi
    era_ammin = a.is_ammin
    db.delete(a)
    db.commit()
    if era_pi or era_ammin:
        background_tasks.add_task(_trigger_sync_progetti)
    return {"data": {"deleted": True}}


# ─── Budget per WP ────────────────────────────────────────────────────────────

@router.post("/{id}/budget/wp")
def salva_budget_wp(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from collections import defaultdict
    p = _get_or_404(id, db)
    if not p.gestione_per_wp:
        raise HTTPException(status_code=422, detail={"error": {"code": "GESTIONE_WP_DISABILITATA",
                                                                "message": "Il progetto non usa gestione per WP"}})
    voci_input = body.get("voci", [])  # [{voce_id, wp_id, importo_previsto}]

    # Ripartizione per voce deve uguagliare il totale di progetto
    bv_proj = {str(bv.voce_id): float(bv.importo_previsto)
               for bv in db.query(BudgetVoce).filter(BudgetVoce.progetto_id == id, BudgetVoce.wp_id.is_(None)).all()}
    totale_per_voce: dict = defaultdict(float)
    for v in voci_input:
        totale_per_voce[str(v["voce_id"])] += float(v.get("importo_previsto", 0))
    for voce_id, totale in totale_per_voce.items():
        totale_proj = bv_proj.get(voce_id, 0.0)
        if abs(totale - totale_proj) > 0.01:
            raise HTTPException(status_code=422, detail={"error": {
                "code": "BUDGET_WP_NON_BILANCIA",
                "message": f"La somma WP ({totale:,.2f}€) non corrisponde al totale di progetto ({totale_proj:,.2f}€)"
            }})

    db.query(BudgetVoce).filter(BudgetVoce.progetto_id == id, BudgetVoce.wp_id.isnot(None)).delete(synchronize_session=False)
    for v in voci_input:
        if float(v.get("importo_previsto", 0)) > 0:
            db.add(BudgetVoce(progetto_id=id, voce_id=v["voce_id"], wp_id=v["wp_id"],
                              importo_previsto=v["importo_previsto"], importo_rendicontato=0))
    db.commit()
    return {"data": {"saved": True}}


# ─── Allocazioni per WP (batch) ───────────────────────────────────────────────

@router.post("/{id}/allocazioni/wp")
def salva_allocazioni_wp(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from collections import defaultdict
    p = _get_or_404(id, db)
    if not p.gestione_per_wp:
        raise HTTPException(status_code=422, detail={"error": {"code": "GESTIONE_WP_DISABILITATA",
                                                                "message": "Il progetto non usa gestione per WP"}})
    alloc_input = body.get("allocazioni", [])  # [{persona_id, wp_id, ore_assegnate}]

    # Ripartizione per persona deve uguagliare il totale di progetto
    alloc_proj = {str(a.persona_id): a
                  for a in db.query(Allocazione).filter(Allocazione.progetto_id == id, Allocazione.wp_id.is_(None)).all()}
    ore_per_persona: dict = defaultdict(float)
    for a in alloc_input:
        ore_per_persona[str(a["persona_id"])] += float(a.get("ore_assegnate", 0))
    for persona_id, ore in ore_per_persona.items():
        ore_proj = float(alloc_proj[persona_id].ore_assegnate) if persona_id in alloc_proj else 0.0
        if abs(ore - ore_proj) > 0.01:
            raise HTTPException(status_code=422, detail={"error": {
                "code": "ORE_WP_NON_BILANCIATE",
                "message": f"La somma ore WP ({ore}) non corrisponde alle ore totali ({ore_proj}) per la persona"
            }})

    db.query(Allocazione).filter(Allocazione.progetto_id == id, Allocazione.wp_id.isnot(None)).delete(synchronize_session=False)
    for a in alloc_input:
        if float(a.get("ore_assegnate", 0)) > 0:
            proj_a = alloc_proj.get(str(a["persona_id"]))
            db.add(Allocazione(
                progetto_id=id, persona_id=a["persona_id"], wp_id=a["wp_id"],
                ore_assegnate=a["ore_assegnate"],
                data_inizio=proj_a.data_inizio if proj_a else p.data_inizio,
                data_fine=proj_a.data_fine if proj_a else p.data_fine,
                is_pi=False, is_ammin=False,
            ))
    db.commit()
    return {"data": {"saved": True}}


# ─── Helper ───────────────────────────────────────────────────────────────────

def _get_or_404(id: str, db: Session) -> Progetto:
    p = db.query(Progetto).filter(Progetto.id == id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Progetto non trovato"}})
    return p


def _transizione(id: str, azione: str, db: Session):
    p = _get_or_404(id, db)
    nuovo_stato = TRANSIZIONI.get(p.stato, {}).get(azione)
    if not nuovo_stato:
        raise HTTPException(status_code=409,
                            detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
                                              "message": f"Impossibile '{azione}' da stato '{p.stato}'"}})
    p.stato = nuovo_stato
    db.commit()
    db.refresh(p)
    return {"data": progetto_dict(p)}


# ─── Validazioni helper ───────────────────────────────────────────────────────

def _valida_date_nel_progetto(progetto: Progetto, data_inizio: date, data_fine: date, entita: str):
    """Verifica che le date siano comprese nel periodo del progetto."""
    if data_inizio < progetto.data_inizio:
        raise HTTPException(
            status_code=422,
            detail={"error": {"code": "DATE_FUORI_PROGETTO",
                               "message": f"{entita}: la data di inizio ({data_inizio.strftime('%d/%m/%Y')}) è precedente all'inizio del progetto ({progetto.data_inizio.strftime('%d/%m/%Y')})"}},
        )
    if data_fine > progetto.data_fine:
        raise HTTPException(
            status_code=422,
            detail={"error": {"code": "DATE_FUORI_PROGETTO",
                               "message": f"{entita}: la data di fine ({data_fine.strftime('%d/%m/%Y')}) è successiva alla fine del progetto ({progetto.data_fine.strftime('%d/%m/%Y')})"}},
        )
    if data_inizio > data_fine:
        raise HTTPException(
            status_code=422,
            detail={"error": {"code": "DATE_NON_VALIDE",
                               "message": f"{entita}: la data di inizio deve essere precedente alla data di fine"}},
        )


# ─── Spese ────────────────────────────────────────────────────────────────────

@router.get("/{id}/spese")
def lista_spese(
    id: str,
    voce_id: str = Query(None),
    sal_id: str = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.budget import Spesa, VoceDiCosto
    _get_or_404(id, db)
    q = db.query(Spesa).filter(Spesa.progetto_id == id)
    if voce_id:
        q = q.filter(Spesa.voce_id == voce_id)
    if sal_id:
        q = q.filter(Spesa.sal_id == sal_id)
    q = q.order_by(Spesa.data.desc())
    total = q.count()
    items = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "data": [_spesa_dict(s) for s in items],
        "meta": {"total": total, "page": page, "page_size": page_size,
                 "total_pages": max(1, (total + page_size - 1) // page_size)},
    }


@router.post("/{id}/spese")
def registra_spesa(
    id: str,
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    if utente.ruolo == "amministrativo":
        pass  # sempre autorizzato
    else:
        from app.models.personale import Allocazione
        e_pi = db.query(Allocazione).filter(
            Allocazione.persona_id == utente.id,
            Allocazione.progetto_id == id,
            Allocazione.is_pi == True,
        ).first()
        if not e_pi:
            raise HTTPException(status_code=403, detail={"error": {
                "code": "FORBIDDEN", "message": "Solo l'amministrativo o il PI del progetto possono registrare spese"}})
    from app.models.budget import Spesa, BudgetVoce
    import uuid as _uuid
    p = _get_or_404(id, db)
    impegno_id = body.get("impegno_id")
    spesa = Spesa(
        id=_uuid.uuid4(),
        progetto_id=id,
        voce_id=body.get("voce_id"),
        wp_id=body.get("wp_id"),
        persona_id=body.get("persona_id"),
        sal_id=body.get("sal_id"),
        impegno_id=impegno_id,
        spesa_origine_id=body.get("spesa_origine_id"),
        importo=body.get("importo", 0),
        data=body.get("data"),
        data_documento=body.get("data_documento"),
        numero_documento=body.get("numero_documento"),
        descrizione=body.get("descrizione"),
        stato="registrata",
        rendicontata=False,
    )
    db.add(spesa)
    # Recupera BudgetVoce per aggiornare l'impegno (se presente)
    bv = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == id,
        BudgetVoce.voce_id == spesa.voce_id,
    ).first()
    # Stabilizza l'impegno collegato (opzione A: rimuove l'intero impegno dal contabilizzato)
    if impegno_id:
        imp = db.query(Impegno).filter(Impegno.id == impegno_id).first()
        if imp:
            # Blocca se l'impegno è già stabilizzato
            gia_usato = db.query(Spesa).filter(
                Spesa.impegno_id == impegno_id,
                Spesa.id != spesa.id,
            ).first()
            if gia_usato:
                db.rollback()
                raise HTTPException(status_code=400, detail={"error": {
                    "code": "IMPEGNO_GIA_STABILIZZATO",
                    "message": "L'impegno selezionato è già collegato a un'altra spesa",
                }})
            if bv:
                bv.importo_impegnato = max(0, float(bv.importo_impegnato) - float(imp.importo))
    db.commit()
    db.refresh(spesa)
    return {"data": _spesa_dict(spesa)}


@router.post("/spese/{id}/annulla")
def annulla_spesa(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    from app.models.budget import Spesa, BudgetVoce
    spesa = db.query(Spesa).filter(Spesa.id == id).first()
    if not spesa:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Spesa non trovata"}})
    # Storna importo_rendicontato
    bv = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == spesa.progetto_id,
        BudgetVoce.voce_id == spesa.voce_id,
    ).first()
    # Ripristina l'impegno collegato (destabilizza)
    if spesa.impegno_id:
        imp = db.query(Impegno).filter(Impegno.id == spesa.impegno_id).first()
        if imp and bv:
            bv.importo_impegnato = float(bv.importo_impegnato) + float(imp.importo)
    db.delete(spesa)
    db.commit()
    return {"data": {"deleted": True}}


@router.post("/spese/{id}/allegato")
def upload_allegato(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.budget import Spesa
    spesa = db.query(Spesa).filter(Spesa.id == id).first()
    if not spesa:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Spesa non trovata"}})
    return {"data": {"allegato_path": None, "message": "Upload non ancora implementato"}}


def _spesa_dict(s) -> dict:
    from app.models.budget import VoceDiCosto
    return {
        "id": str(s.id),
        "progetto_id": str(s.progetto_id),
        "voce_id": str(s.voce_id) if s.voce_id else None,
        "wp_id": str(s.wp_id) if s.wp_id else None,
        "persona_id": str(s.persona_id) if s.persona_id else None,
        "sal_id": str(s.sal_id) if s.sal_id else None,
        "impegno_id": str(s.impegno_id) if s.impegno_id else None,
        "spesa_origine_id": str(s.spesa_origine_id) if s.spesa_origine_id else None,
        "importo": float(s.importo),
        "data": str(s.data) if s.data else None,
        "data_documento": str(s.data_documento) if s.data_documento else None,
        "numero_documento": s.numero_documento,
        "descrizione": s.descrizione,
        "stato": s.stato,
        "rendicontata": s.rendicontata,
        "allegato_path": s.allegato_path,
    }


# ─── Impegni ──────────────────────────────────────────────────────────────────

def _impegno_dict(i: "Impegno", stabilizzato: bool = False) -> dict:
    return {
        "id": str(i.id),
        "progetto_id": str(i.progetto_id),
        "voce_id": str(i.voce_id),
        "wp_id": str(i.wp_id) if i.wp_id else None,
        "voce": {"codice": i.voce.codice, "descrizione": i.voce.descrizione} if i.voce else None,
        "data": str(i.data) if i.data else None,
        "descrizione": i.descrizione,
        "importo": float(i.importo),
        "stabilizzato": stabilizzato,
        "created_at": str(i.created_at) if i.created_at else None,
    }


def _impegni_stabilizzati_ids(progetto_id: str, db: Session) -> set:
    """Restituisce gli ID degli impegni già collegati a una spesa (stabilizzati)."""
    rows = db.query(Spesa.impegno_id).filter(
        Spesa.impegno_id.isnot(None),
        Spesa.progetto_id == progetto_id,
    ).all()
    return {str(r[0]) for r in rows}


def _get_bv_or_404(progetto_id: str, voce_id: str, db: Session) -> "BudgetVoce":
    bv = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == progetto_id,
        BudgetVoce.voce_id == voce_id,
    ).first()
    if not bv:
        raise HTTPException(status_code=404, detail={"error": {"code": "VOCE_NON_TROVATA", "message": "Voce di costo non presente nel budget del progetto"}})
    return bv


@router.get("/{id}/impegni")
def lista_impegni(
    id: str,
    voce_id: str = Query(None),
    solo_disponibili: bool = Query(False),
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    _get_or_404(id, db)
    q = db.query(Impegno).filter(Impegno.progetto_id == id)
    if voce_id:
        q = q.filter(Impegno.voce_id == voce_id)
    stabilizzati = _impegni_stabilizzati_ids(id, db)
    if solo_disponibili:
        q = q.filter(~Impegno.id.in_([s for s in stabilizzati]))
    impegni = q.order_by(Impegno.data.desc()).all()
    return {"data": [_impegno_dict(i, stabilizzato=str(i.id) in stabilizzati) for i in impegni]}


@router.post("/{id}/impegni")
def crea_impegno(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    _get_or_404(id, db)
    voce_id = body.get("voce_id")
    importo = float(body.get("importo", 0))
    bv = _get_bv_or_404(id, voce_id, db)
    speso = float(db.query(func.sum(Spesa.importo)).filter(
        Spesa.progetto_id == id, Spesa.voce_id == voce_id, Spesa.stato == "registrata"
    ).scalar() or 0)
    disponibile = float(bv.importo_previsto) - float(bv.importo_impegnato) - speso
    if importo > disponibile:
        raise HTTPException(status_code=422, detail={"error": {
            "code": "CAPIENZA_INSUFFICIENTE",
            "message": f"Capienza insufficiente sulla voce. Disponibile: {disponibile:,.2f} €",
        }})
    impegno = Impegno(
        progetto_id=id,
        voce_id=voce_id,
        wp_id=body.get("wp_id"),
        data=body.get("data"),
        descrizione=body.get("descrizione", ""),
        importo=importo,
        created_by=utente.id,
    )
    db.add(impegno)
    bv.importo_impegnato = float(bv.importo_impegnato) + importo
    db.commit()
    db.refresh(impegno)
    return {"data": _impegno_dict(impegno)}


@router.put("/impegni/{impegno_id}")
def modifica_impegno(impegno_id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    impegno = db.query(Impegno).filter(Impegno.id == impegno_id).first()
    if not impegno:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Impegno non trovato"}})
    if db.query(Spesa).filter(Spesa.impegno_id == impegno_id).first():
        raise HTTPException(status_code=400, detail={"error": {"code": "IMPEGNO_STABILIZZATO",
                                                                "message": "Impossibile modificare un impegno già collegato a una spesa"}})
    nuovo_importo = float(body.get("importo", impegno.importo))
    delta = nuovo_importo - float(impegno.importo)
    if delta != 0:
        bv = _get_bv_or_404(str(impegno.progetto_id), str(impegno.voce_id), db)
        speso = float(db.query(func.sum(Spesa.importo)).filter(
            Spesa.progetto_id == impegno.progetto_id, Spesa.voce_id == impegno.voce_id, Spesa.stato == "registrata"
        ).scalar() or 0)
        disponibile = float(bv.importo_previsto) - float(bv.importo_impegnato) - speso
        if delta > disponibile:
            raise HTTPException(status_code=422, detail={"error": {
                "code": "CAPIENZA_INSUFFICIENTE",
                "message": f"Capienza insufficiente sulla voce. Disponibile aggiuntivo: {disponibile:,.2f} €",
            }})
        bv.importo_impegnato = float(bv.importo_impegnato) + delta
    if "data" in body:
        impegno.data = body["data"]
    if "descrizione" in body:
        impegno.descrizione = body["descrizione"]
    impegno.importo = nuovo_importo
    db.commit()
    db.refresh(impegno)
    return {"data": _impegno_dict(impegno)}


@router.delete("/impegni/{impegno_id}")
def elimina_impegno(impegno_id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    impegno = db.query(Impegno).filter(Impegno.id == impegno_id).first()
    if not impegno:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Impegno non trovato"}})
    if db.query(Spesa).filter(Spesa.impegno_id == impegno_id).first():
        raise HTTPException(status_code=400, detail={"error": {"code": "IMPEGNO_STABILIZZATO",
                                                                "message": "Impossibile eliminare un impegno già collegato a una spesa"}})
    bv = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == impegno.progetto_id,
        BudgetVoce.voce_id == impegno.voce_id,
    ).first()
    if bv:
        bv.importo_impegnato = max(0, float(bv.importo_impegnato) - float(impegno.importo))
    db.delete(impegno)
    db.commit()
    return {"data": {"deleted": True}}


# ─── Documenti ────────────────────────────────────────────────────────────────

@router.get("/{id}/documenti")
def lista_documenti(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.documento import DocumentoProgetto
    _get_or_404(id, db)
    docs = db.query(DocumentoProgetto).filter(
        DocumentoProgetto.progetto_id == id
    ).order_by(DocumentoProgetto.uploaded_at.desc()).all()
    return {"data": [_doc_dict(d) for d in docs]}


@router.post("/{id}/documenti")
async def upload_documento(
    id: str,
    file: UploadFile = File(...),
    tipo_documento: str = "altro",
    versione: str = None,
    descrizione: str = None,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    if utente.ruolo not in ("amministrativo", "pi"):
        raise HTTPException(status_code=403, detail={"error": {
            "code": "FORBIDDEN", "message": "Solo amministrativi e PI possono caricare documenti"}})
    from app.models.documento import DocumentoProgetto
    import os, uuid as _uuid
    _progetto = _get_or_404(id, db)
    from app.services.storage import progetto_dir, upload_filename
    upload_dir = progetto_dir(_progetto.codice, "documenti")
    os.makedirs(upload_dir, exist_ok=True)

    doc_id = _uuid.uuid4()
    ext = os.path.splitext(file.filename)[1] if file.filename else ""
    dest = os.path.join(upload_dir, upload_filename(file.filename or f"doc{ext}", str(doc_id)))

    content_bytes = await file.read()
    with open(dest, "wb") as f_out:
        f_out.write(content_bytes)

    doc = DocumentoProgetto(
        id=doc_id,
        progetto_id=id,
        tipo_documento=tipo_documento,
        nome_file=file.filename or f"{doc_id}{ext}",
        path_file=dest,
        versione=versione,
        descrizione=descrizione,
        uploaded_by=utente.id,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return {"data": _doc_dict(doc)}


@router.delete("/documenti/{doc_id}")
def elimina_documento(
    doc_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.documento import DocumentoProgetto
    import os
    doc = db.query(DocumentoProgetto).filter(DocumentoProgetto.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Documento non trovato"}})
    if os.path.exists(doc.path_file):
        os.remove(doc.path_file)
    db.delete(doc)
    db.commit()
    return {"data": {"deleted": True}}


@router.get("/documenti/{doc_id}/download")
def download_documento(
    doc_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.documento import DocumentoProgetto
    from fastapi.responses import FileResponse
    import os
    doc = db.query(DocumentoProgetto).filter(DocumentoProgetto.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Documento non trovato"}})
    if not os.path.exists(doc.path_file):
        raise HTTPException(status_code=404, detail={"error": {"code": "FILE_NON_TROVATO", "message": "File non trovato sul server"}})
    return FileResponse(doc.path_file, filename=doc.nome_file)


@router.patch("/documenti/{doc_id}")
def aggiorna_documento(
    doc_id: str,
    body: dict = Body(...),
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.documento import DocumentoProgetto
    doc = db.query(DocumentoProgetto).filter(DocumentoProgetto.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Documento non trovato"}})

    if "descrizione" in body:
        doc.descrizione = body.get("descrizione")
    if "tipo_documento" in body:
        doc.tipo_documento = body.get("tipo_documento")

    db.commit()
    return {"data": _doc_dict(doc)}


def _doc_dict(d) -> dict:
    return {
        "id": str(d.id),
        "progetto_id": str(d.progetto_id),
        "tipo_documento": d.tipo_documento,
        "nome_file": d.nome_file,
        "versione": d.versione,
        "descrizione": d.descrizione,
        "uploaded_at": d.uploaded_at.isoformat() if d.uploaded_at else None,
        "uploaded_by": str(d.uploaded_by) if d.uploaded_by else None,
    }


@router.delete("/{id}")
def elimina_progetto(
    id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    from app.models.partner import ProgettoPartner
    from app.models.struttura import WorkPackage
    from app.models.personale import Allocazione
    from app.models.budget import BudgetVoce, Spesa, Sal
    from app.models.documento import DocumentoProgetto
    from app.models.timesheet import TimesheetTestata

    p = _get_or_404(id, db)
    if p.stato != "bozza" and utente.ruolo != "superadmin":
        raise HTTPException(status_code=409, detail={"error": {
            "code": "PROGETTO_NON_ELIMINABILE",
            "message": "Solo i progetti in stato bozza possono essere eliminati"
        }})

    # Elimina manualmente tutte le righe correlate
    db.query(ProgettoPartner).filter(ProgettoPartner.progetto_id == p.id).delete()
    db.query(Allocazione).filter(Allocazione.progetto_id == p.id).delete()
    db.query(BudgetVoce).filter(BudgetVoce.progetto_id == p.id).delete()
    db.query(Spesa).filter(Spesa.progetto_id == p.id).delete()
    db.query(Sal).filter(Sal.progetto_id == p.id).delete()
    db.query(DocumentoProgetto).filter(DocumentoProgetto.progetto_id == p.id).delete()

    # WP prima dei timesheet (i timesheet referenziano i WP)
    wp_ids = [str(w.id) for w in db.query(WorkPackage).filter(WorkPackage.progetto_id == p.id).all()]
    db.query(WorkPackage).filter(WorkPackage.progetto_id == p.id).delete()

    db.query(TimesheetTestata).filter(TimesheetTestata.progetto_id == p.id).delete()

    db.delete(p)
    db.commit()
    background_tasks.add_task(_notifica_sync_missioni)
    return {"data": {"deleted": True}}


# ─── Milestone ────────────────────────────────────────────────────────────────

@router.get("/{id}/milestone")
def lista_milestone(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    from app.models.struttura import Milestone
    _get_or_404(id, db)
    items = db.query(Milestone).filter(Milestone.progetto_id == id).order_by(Milestone.data_prevista).all()
    return {"data": [_milestone_dict(m) for m in items]}

@router.post("/{id}/milestone")
def crea_milestone(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Milestone
    import uuid as _uuid
    _get_or_404(id, db)
    m = Milestone(id=_uuid.uuid4(), progetto_id=id, **{k: v for k, v in body.items() if k != 'progetto_id' and hasattr(Milestone, k)})
    db.add(m); db.commit(); db.refresh(m)
    return {"data": _milestone_dict(m)}

@router.patch("/{id}/milestone/{ms_id}")
def aggiorna_milestone(id: str, ms_id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Milestone
    m = db.query(Milestone).filter(Milestone.id == ms_id, Milestone.progetto_id == id).first()
    if not m: raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Milestone non trovata"}})
    for k, v in body.items():
        if hasattr(Milestone, k) and k not in ('id', 'progetto_id'): setattr(m, k, v)
    db.commit(); db.refresh(m)
    return {"data": _milestone_dict(m)}

@router.delete("/{id}/milestone/{ms_id}")
def elimina_milestone(id: str, ms_id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Milestone
    m = db.query(Milestone).filter(Milestone.id == ms_id, Milestone.progetto_id == id).first()
    if not m: raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Milestone non trovata"}})
    db.delete(m); db.commit()
    return {"data": {"deleted": True}}

def _milestone_dict(m) -> dict:
    return {"id": str(m.id), "progetto_id": str(m.progetto_id), "wp_id": str(m.wp_id) if m.wp_id else None,
            "codice": m.codice, "titolo": m.titolo, "data_prevista": str(m.data_prevista) if m.data_prevista else None,
            "data_effettiva": str(m.data_effettiva) if m.data_effettiva else None, "stato": m.stato}


# ─── Deliverable ──────────────────────────────────────────────────────────────

@router.get("/{id}/deliverable")
def lista_deliverable(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    from app.models.struttura import Deliverable
    _get_or_404(id, db)
    items = db.query(Deliverable).filter(Deliverable.progetto_id == id).order_by(Deliverable.data_scadenza).all()
    return {"data": [_deliverable_dict(d) for d in items]}

@router.post("/{id}/deliverable")
def crea_deliverable(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Deliverable
    import uuid as _uuid
    _get_or_404(id, db)
    d = Deliverable(id=_uuid.uuid4(), progetto_id=id, **{k: v for k, v in body.items() if k != 'progetto_id' and hasattr(Deliverable, k)})
    db.add(d); db.commit(); db.refresh(d)
    return {"data": _deliverable_dict(d)}

@router.patch("/{id}/deliverable/{del_id}")
def aggiorna_deliverable(id: str, del_id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Deliverable
    d = db.query(Deliverable).filter(Deliverable.id == del_id, Deliverable.progetto_id == id).first()
    if not d: raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Deliverable non trovato"}})
    for k, v in body.items():
        if hasattr(Deliverable, k) and k not in ('id', 'progetto_id'): setattr(d, k, v)
    db.commit(); db.refresh(d)
    return {"data": _deliverable_dict(d)}

@router.delete("/{id}/deliverable/{del_id}")
def elimina_deliverable(id: str, del_id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    from app.models.struttura import Deliverable
    d = db.query(Deliverable).filter(Deliverable.id == del_id, Deliverable.progetto_id == id).first()
    if not d: raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Deliverable non trovato"}})
    db.delete(d); db.commit()
    return {"data": {"deleted": True}}

def _deliverable_dict(d) -> dict:
    return {"id": str(d.id), "progetto_id": str(d.progetto_id), "wp_id": str(d.wp_id) if d.wp_id else None,
            "codice": d.codice, "titolo": d.titolo, "tipo": d.tipo,
            "data_scadenza": str(d.data_scadenza) if d.data_scadenza else None,
            "data_consegna": str(d.data_consegna) if d.data_consegna else None, "stato": d.stato}


# ─── Notifiche ai partecipanti ────────────────────────────────────────────────

@router.post("/{id}/notifica-partecipanti")
def invia_notifica_partecipanti(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    p = _get_or_404(id, db)

    if str(p.amministrativo_id) != str(utente.id) and utente.ruolo != "superadmin":
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN", "message": "Non sei amministrativo di questo progetto"}})

    titolo = body.get("titolo", "Notifica del progetto")
    messaggio = body.get("messaggio", "")

    if not messaggio.strip():
        raise HTTPException(status_code=400, detail={"error": {"code": "BAD_REQUEST", "message": "Messaggio non può essere vuoto"}})

    allocazioni = db.query(Allocazione).filter(Allocazione.progetto_id == id).all()
    persone_ids = set(str(a.persona_id) for a in allocazioni)
    persone_ids.discard(str(utente.id))

    notifiche_create = []
    for persona_id in persone_ids:
        n = crea_notifica(
            db,
            persona_id=persona_id,
            tipo="messaggio_partecipanti",
            titolo=titolo,
            messaggio=messaggio,
        )
        db.flush()
        notifiche_create.append(str(n.id))

    db.commit()
    return {"data": {"notifiche_create": len(notifiche_create)}}


# ─── Helpers report ───────────────────────────────────────────────────────────

def _check_report_permission(p: Progetto, utente: Persona):
    if utente.ruolo == "superadmin":
        return
    if p.amministrativo_id and str(p.amministrativo_id) == str(utente.id):
        return
    raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN",
        "message": "Solo l'amministrativo del progetto può generare il report"}})


def _raccogli_dati_report(id: str, db: Session):
    p = db.query(Progetto).filter(Progetto.id == id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Progetto non trovato"}})

    pi_alloc = db.query(Allocazione).filter(Allocazione.progetto_id == id, Allocazione.is_pi == True).first()
    pi = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first() if pi_alloc else None
    amm = db.query(Persona).filter(Persona.id == p.amministrativo_id).first() if p.amministrativo_id else None

    budget_voci = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == id, BudgetVoce.wp_id.is_(None)
    ).all()
    voci_map = {str(v.id): v for v in db.query(VoceDiCosto).all()}

    spese = db.query(Spesa).filter(Spesa.progetto_id == id, Spesa.stato == "registrata").order_by(Spesa.data).all()

    timesheet = (db.query(TimesheetTestata)
                 .filter(TimesheetTestata.progetto_id == id, TimesheetTestata.stato == "approvato")
                 .order_by(TimesheetTestata.anno, TimesheetTestata.mese).all())
    persone_map = {str(pe.id): pe for pe in db.query(Persona).all()}

    sal_list = db.query(Sal).filter(Sal.progetto_id == id).order_by(Sal.numero).all()

    partner_rows = db.query(ProgettoPartner).filter(ProgettoPartner.progetto_id == id).all()
    partner_map = {str(pa.id): pa for pa in db.query(Partner).all()}

    return p, pi, amm, budget_voci, voci_map, spese, timesheet, persone_map, sal_list, partner_rows, partner_map


# ─── Riepilogo dashboard Excel ────────────────────────────────────────────────

@router.get("/{id}/riepilogo-dashboard/xlsx")
def riepilogo_dashboard_xlsx(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    p = _get_or_404(id, db)

    # ── dati ──────────────────────────────────────────────────────────────────
    # PI
    pi_alloc = db.query(Allocazione).filter(Allocazione.progetto_id == id, Allocazione.is_pi == True).first()
    pi_nome = None
    if pi_alloc:
        pi_p = db.query(Persona).filter(Persona.id == pi_alloc.persona_id).first()
        pi_nome = f"{pi_p.nome} {pi_p.cognome}" if pi_p else None

    # budget voci con speso (solo voci di progetto, non sotto-voci WP)
    budget_voci = db.query(BudgetVoce).filter(
        BudgetVoce.progetto_id == id, BudgetVoce.wp_id.is_(None)
    ).all()
    spese_per_voce = dict(
        db.query(Spesa.voce_id, func.sum(Spesa.importo))
        .filter(Spesa.progetto_id == id, Spesa.stato == "registrata")
        .group_by(Spesa.voce_id).all()
    )

    # spese
    spese = db.query(Spesa).filter(Spesa.progetto_id == id, Spesa.stato == "registrata").order_by(Spesa.data).all()
    voci_map = {str(v.id): v for v in db.query(VoceDiCosto).all()}

    # impegni
    impegni = db.query(Impegno).filter(Impegno.progetto_id == id).order_by(Impegno.data).all()

    # SAL
    sal_list = db.query(Sal).filter(Sal.progetto_id == id).order_by(Sal.numero).all()

    # KPI aggregati
    tot_pianificato = sum(float(v.importo_previsto) for v in budget_voci)
    tot_impegnato   = sum(float(v.importo_impegnato) for v in budget_voci)
    tot_speso       = sum(float(spese_per_voce.get(v.voce_id, 0)) for v in budget_voci)
    tot_rendicontato = sum(float(v.importo_rendicontato) for v in budget_voci)
    tot_disponibile = max(0, tot_pianificato - tot_impegnato - tot_speso)

    oggi = date.today()
    if p.data_inizio and p.data_fine:
        durata = (p.data_fine - p.data_inizio).days
        trascorsi = (oggi - p.data_inizio).days
        pct_tempo = round(max(0, min(trascorsi / durata * 100, 100)), 1) if durata > 0 else 0
    else:
        pct_tempo = 0

    # ── workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    ROSSO  = "FFC5174E"
    BIANCO = "FFFFFFFF"
    GRIGIO = "FFF5F5F5"
    VIOLA  = "FF722ED1"
    TEAL   = "FF13C2C2"
    ORANGE = "FFE8863A"

    def hdr(ws, row, n_cols, colore=ROSSO):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(bold=True, color=BIANCO)
            cell.fill = PatternFill("solid", fgColor=colore)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_w(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 55)

    def eur(v):
        return round(float(v or 0), 2)

    def row_grigio(ws, row, n_cols):
        for c in range(1, n_cols + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFE8E8E8")

    # ── Sheet 1: Riepilogo ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Riepilogo"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 50

    def kv(label, value, bold_val=False):
        r = ws1.max_row + 1
        ws1.cell(r, 1, label).font = Font(bold=True, color="FF555555")
        c = ws1.cell(r, 2, value)
        if bold_val:
            c.font = Font(bold=True)

    # Titolo
    ws1.merge_cells("A1:B1")
    t = ws1.cell(1, 1, f"RIEPILOGO PROGETTO — {p.acronimo}")
    t.font = Font(bold=True, size=14, color=BIANCO)
    t.fill = PatternFill("solid", fgColor=ROSSO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.append([])
    ws1.append(["ANAGRAFICA", ""])
    hdr(ws1, ws1.max_row, 2, "FF888888")
    kv("Codice", p.codice)
    kv("Acronimo", p.acronimo)
    kv("Titolo", p.titolo)
    kv("Tipo", str(p.tipo) if p.tipo else "—")
    kv("Riferimento bando", p.riferimento_bando or "—")
    kv("PI", pi_nome or "—")
    kv("Data inizio", str(p.data_inizio) if p.data_inizio else "—")
    kv("Data fine", str(p.data_fine) if p.data_fine else "—")
    kv("Stato", p.stato)
    ws1.append([])
    ws1.append(["IMPORTI", ""])
    hdr(ws1, ws1.max_row, 2, "FF888888")
    kv("Costo totale progetto", eur(p.costo_totale), True)
    kv("Importo finanziato", eur(p.importo_finanziato), True)
    kv("Cofinanziamento", eur(float(p.costo_totale or 0) - float(p.importo_finanziato or 0)), True)
    ws1.append([])
    ws1.append(["KPI FINANZIARI", ""])
    hdr(ws1, ws1.max_row, 2, "FF888888")
    kv("Budget pianificato", eur(tot_pianificato))
    kv("Rendicontato", eur(tot_rendicontato))
    kv("Spese totali", eur(tot_speso))
    kv("Impegnato totale", eur(tot_impegnato))
    kv("Disponibile", eur(tot_disponibile))
    kv("Avanzamento temporale", f"{pct_tempo}%")

    # ── Sheet 2: Budget per voce ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Budget per voce")
    headers2 = ["Voce di costo", "Pianificato €", "Impegnato €", "Speso €", "Rendicontato €", "Disponibile €", "% Utilizzato"]
    ws2.append(headers2)
    hdr(ws2, 1, len(headers2))
    for i, v in enumerate(budget_voci):
        speso_v = float(spese_per_voce.get(v.voce_id, 0))
        disp_v  = max(0, float(v.importo_previsto) - float(v.importo_impegnato) - speso_v)
        pct_v   = round(float(v.importo_rendicontato) / float(v.importo_previsto) * 100, 1) if float(v.importo_previsto) > 0 else 0
        row = [
            v.voce.descrizione if v.voce else str(v.voce_id),
            eur(v.importo_previsto),
            eur(v.importo_impegnato),
            eur(speso_v),
            eur(v.importo_rendicontato),
            eur(disp_v),
            f"{pct_v}%",
        ]
        ws2.append(row)
        if i % 2 == 1:
            row_grigio(ws2, ws2.max_row, len(headers2))
    # totali
    tr = ws2.max_row + 1
    ws2.append(["TOTALE", eur(tot_pianificato), eur(tot_impegnato), eur(tot_speso), eur(tot_rendicontato), eur(tot_disponibile), ""])
    for c in range(1, len(headers2) + 1):
        ws2.cell(tr, c).font = Font(bold=True)
        ws2.cell(tr, c).fill = PatternFill("solid", fgColor="FFE0E0E0")
    auto_w(ws2)

    # ── Sheet 3: Impegni ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Impegni")
    headers3 = ["Data", "Voce di costo", "Descrizione", "Importo €"]
    ws3.append(headers3)
    hdr(ws3, 1, len(headers3), VIOLA)
    for i, imp in enumerate(impegni):
        voce_desc = imp.voce.descrizione if imp.voce else str(imp.voce_id)
        ws3.append([str(imp.data), voce_desc, imp.descrizione, eur(imp.importo)])
        if i % 2 == 1:
            row_grigio(ws3, ws3.max_row, len(headers3))
    if impegni:
        tr3 = ws3.max_row + 1
        ws3.append(["", "", "TOTALE", eur(sum(float(i.importo) for i in impegni))])
        ws3.cell(tr3, 3).font = Font(bold=True)
        ws3.cell(tr3, 4).font = Font(bold=True)
    auto_w(ws3)

    # ── Sheet 4: Spese ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Spese")
    headers4 = ["Data", "Voce di costo", "N° documento", "Descrizione", "Importo €"]
    ws4.append(headers4)
    hdr(ws4, 1, len(headers4), ORANGE)
    for i, s in enumerate(spese):
        voce_desc = voci_map.get(str(s.voce_id), VoceDiCosto()).descrizione if s.voce_id else "—"
        ws4.append([str(s.data), voce_desc, s.numero_documento or "—", s.descrizione or "—", eur(s.importo)])
        if i % 2 == 1:
            row_grigio(ws4, ws4.max_row, len(headers4))
    if spese:
        tr4 = ws4.max_row + 1
        ws4.append(["", "", "", "TOTALE", eur(sum(float(s.importo) for s in spese))])
        ws4.cell(tr4, 4).font = Font(bold=True)
        ws4.cell(tr4, 5).font = Font(bold=True)
    auto_w(ws4)

    # ── Sheet 5: SAL ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("SAL")
    headers5 = ["N° SAL", "Data inizio", "Data fine", "Stato", "Importo tranche €", "Importo erogato €", "Scadenza rendiconto"]
    ws5.append(headers5)
    hdr(ws5, 1, len(headers5), TEAL)
    for i, s in enumerate(sal_list):
        ws5.append([
            s.numero, str(s.data_inizio), str(s.data_fine), s.stato,
            eur(s.importo_tranche) if s.importo_tranche else "—",
            eur(s.importo_erogato) if s.importo_erogato else "—",
            str(s.data_scadenza_rendiconto) if s.data_scadenza_rendiconto else "—",
        ])
        if i % 2 == 1:
            row_grigio(ws5, ws5.max_row, len(headers5))
    auto_w(ws5)

    # ── output ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    sigla = (p.acronimo or p.codice or str(p.id)).replace(' ', '_')
    nome = f"Riepilogo_{sigla}_{oggi.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


# ─── Report Excel ─────────────────────────────────────────────────────────────

@router.get("/{id}/report/xlsx")
def report_progetto_xlsx(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    p, pi, amm, budget_voci, voci_map, spese, timesheet, persone_map, sal_list, partner_rows, partner_map = \
        _raccogli_dati_report(id, db)
    _check_report_permission(p, utente)

    wb = openpyxl.Workbook()

    BLU = "FF185FA5"
    BIANCO = "FFFFFFFF"
    GRIGIO = "FFF5F5F5"

    def stile_header(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True, color=BIANCO)
            c.fill = PatternFill("solid", fgColor=BLU)
            c.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    def fmt_eur(v):
        return round(float(v), 2) if v else 0.0

    def fmt_d(d):
        return d.strftime("%d/%m/%Y") if d else ""

    # Sheet 1 — Info progetto
    ws = wb.active
    ws.title = "Progetto"
    rows_info = [
        ("Titolo", p.titolo),
        ("Acronimo", p.acronimo),
        ("Codice", p.codice),
        ("CUP", p.cup or "—"),
        ("Stato", p.stato),
        ("Data inizio", fmt_d(p.data_inizio)),
        ("Data fine", fmt_d(p.data_fine)),
        ("PI", f"{pi.nome} {pi.cognome}" if pi else "—"),
        ("Amministrativo", f"{amm.nome} {amm.cognome}" if amm else "—"),
        ("Costo totale (€)", fmt_eur(p.costo_totale)),
        ("Importo finanziato (€)", fmt_eur(p.importo_finanziato)),
        ("Cofinanziamento (€)", round(fmt_eur(p.costo_totale) - fmt_eur(p.importo_finanziato), 2)),
    ]
    for r, (label, val) in enumerate(rows_info, start=1):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, val)
    auto_width(ws)

    def trunc(s, n=70):
        s = str(s or "")
        return s[:n] + "…" if len(s) > n else s

    # Sheet 2 — Budget
    ws2 = wb.create_sheet("Budget")
    ws2.append(["Voce di costo", "Previsto (€)", "Rendicontato (€)", "Da Rendicontare (€)", "% Utilizzo"])
    stile_header(ws2, 1, 5)
    for bv in budget_voci:
        v = voci_map.get(str(bv.voce_id))
        label = trunc(f"{v.codice} — {v.descrizione}" if v else str(bv.voce_id))
        prev = fmt_eur(bv.importo_previsto)
        rend = fmt_eur(bv.importo_rendicontato)
        pct = round(rend / prev * 100, 1) if prev > 0 else 0
        ws2.append([label, prev, rend, round(prev - rend, 2), pct])
    totale_prev = sum(fmt_eur(bv.importo_previsto) for bv in budget_voci)
    totale_rend = sum(fmt_eur(bv.importo_rendicontato) for bv in budget_voci)
    ws2.append(["TOTALE", totale_prev, totale_rend,
                round(totale_prev - totale_rend, 2),
                round(totale_rend / totale_prev * 100, 1) if totale_prev > 0 else 0])
    r_tot = ws2.max_row
    for col in range(1, 6):
        ws2.cell(r_tot, col).font = Font(bold=True)
    auto_width(ws2)

    # Sheet 3 — Partner
    ws3 = wb.create_sheet("Partner")
    ws3.append(["Nome ente", "Tipo", "Paese", "Ruolo nel progetto", "Budget assegnato (€)"])
    stile_header(ws3, 1, 5)
    for pr in partner_rows:
        pa = partner_map.get(str(pr.partner_id))
        ws3.append([pa.nome if pa else "—", pa.tipo if pa else "—",
                    pa.paese if pa else "—", pr.ruolo,
                    fmt_eur(pr.budget_assegnato) if pr.budget_assegnato else ""])
    auto_width(ws3)

    # Sheet 4 — Spese
    ws4 = wb.create_sheet("Spese")
    ws4.append(["Data", "N° documento", "Voce di costo", "Descrizione", "Importo (€)"])
    stile_header(ws4, 1, 5)
    tot_spese = 0.0
    for s in spese:
        v = voci_map.get(str(s.voce_id))
        voce_label = f"{v.codice} — {v.descrizione}" if v else "—"
        importo = fmt_eur(s.importo)
        tot_spese += importo
        ws4.append([fmt_d(s.data), s.numero_documento or "", voce_label, s.descrizione or "", importo])
    ws4.append(["", "", "", "TOTALE", round(tot_spese, 2)])
    r_tot = ws4.max_row
    for col in range(1, 6):
        ws4.cell(r_tot, col).font = Font(bold=True)
    auto_width(ws4)

    # Sheet 5 — Timesheet
    ws5 = wb.create_sheet("Timesheet")
    ws5.append(["Cognome", "Nome", "Mese", "Anno", "Ore totali", "Costo calcolato (€)"])
    stile_header(ws5, 1, 6)
    tot_ts_costo = 0.0
    for t in timesheet:
        pe = persone_map.get(str(t.persona_id))
        ore = sum(float(c.ore) for r in t.righe for c in r.celle) if t.righe else 0
        costo = sum(float(c.costo_calcolato or 0) for r in t.righe for c in r.celle) if t.righe else 0
        tot_ts_costo += costo
        ws5.append([pe.cognome if pe else "—", pe.nome if pe else "—",
                    t.mese, t.anno, round(ore, 2), round(costo, 2)])
    ws5.append(["", "", "", "", "TOTALE", round(tot_ts_costo, 2)])
    r_tot = ws5.max_row
    for col in range(1, 7):
        ws5.cell(r_tot, col).font = Font(bold=True)
    auto_width(ws5)

    # Sheet 6 — SAL
    ws6 = wb.create_sheet("SAL")
    ws6.append(["N° SAL", "Inizio periodo", "Fine periodo", "Stato", "Importo erogato (€)"])
    stile_header(ws6, 1, 5)
    for s in sal_list:
        ws6.append([s.numero, fmt_d(s.data_inizio), fmt_d(s.data_fine),
                    s.stato, fmt_eur(s.importo_erogato) if s.importo_erogato else ""])
    auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_p = p.acronimo or p.codice or (p.titolo[:30] if p.titolo else "Progetto")
    nome = f"Report_{nome_p}_{datetime.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{nome}"'})


# ─── Report PDF ───────────────────────────────────────────────────────────────

@router.get("/{id}/report/pdf")
def report_progetto_pdf(id: str, db: Session = Depends(get_db), utente: Persona = Depends(tutti_i_ruoli)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    p, pi, amm, budget_voci, voci_map, spese, timesheet, persone_map, sal_list, partner_rows, partner_map = \
        _raccogli_dati_report(id, db)
    _check_report_permission(p, utente)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    BLU = colors.HexColor("#185FA5")
    GRIGIO_CH = colors.HexColor("#F5F5F5")
    BIANCO = colors.white

    styles = getSampleStyleSheet()
    titolo_doc = ParagraphStyle("TitoloDoc", parent=styles["Normal"],
                                fontSize=20, fontName="Helvetica-Bold",
                                textColor=BIANCO, alignment=TA_CENTER)
    sottotitolo = ParagraphStyle("Sottotitolo", parent=styles["Normal"],
                                 fontSize=11, textColor=BIANCO, alignment=TA_CENTER)
    sezione = ParagraphStyle("Sezione", parent=styles["Normal"],
                             fontSize=13, fontName="Helvetica-Bold",
                             textColor=BLU, spaceBefore=16, spaceAfter=6)
    normale = styles["Normal"]
    normale.fontSize = 9

    def fmt_eur(v):
        if not v:
            return "€ 0,00"
        return f"€ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_d(d):
        return d.strftime("%d/%m/%Y") if d else "—"

    def tabella(dati, col_widths, header_row=True, extra_cmds=None):
        t = Table(dati, colWidths=col_widths)
        cmds = [
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BIANCO, GRIGIO_CH]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        if header_row:
            cmds += [
                ("BACKGROUND", (0, 0), (-1, 0), BLU),
                ("TEXTCOLOR", (0, 0), (-1, 0), BIANCO),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        if extra_cmds:
            cmds += extra_cmds
        t.setStyle(TableStyle(cmds))
        return t

    story = []
    W = 17 * cm  # larghezza utile

    # Banner copertina
    banner_data = [[Paragraph(p.acronimo or p.codice or "Progetto", titolo_doc)],
                   [Paragraph(p.titolo or "", sottotitolo)],
                   [Paragraph(f"Codice: {p.codice or '—'}" + (f" | CUP: {p.cup}" if p.cup else ""), sottotitolo)]]
    banner = Table(banner_data, colWidths=[W])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BLU),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
    ]))
    story.append(banner)
    story.append(Spacer(1, 0.4*cm))

    # Info box
    pi_nome = f"{pi.nome} {pi.cognome}" if pi else "—"
    amm_nome = f"{amm.nome} {amm.cognome}" if amm else "—"
    info_data = [
        ["Periodo", f"{fmt_d(p.data_inizio)} → {fmt_d(p.data_fine)}", "Stato", p.stato.upper()],
        ["PI", pi_nome, "Amministrativo", amm_nome],
        ["Costo totale", fmt_eur(p.costo_totale), "Importo finanziato", fmt_eur(p.importo_finanziato)],
        ["Cofinanziamento", fmt_eur(float(p.costo_totale or 0) - float(p.importo_finanziato or 0)),
         "% Finanziato", f"{round(float(p.importo_finanziato or 0) / float(p.costo_totale or 1) * 100, 1)}%"],
    ]
    t_info = Table(info_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
    t_info.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, -1), GRIGIO_CH),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t_info)

    # Budget per voce
    if budget_voci:
        story.append(Paragraph("Budget per voce di costo", sezione))
        bv_data = [["Voce di costo", "Previsto", "Rendicontato", "Da Rendicontare", "%"]]
        tot_prev = tot_rend = 0.0
        for bv in budget_voci:
            v = voci_map.get(str(bv.voce_id))
            label = Paragraph(f"{v.codice} — {v.descrizione}" if v else "—", normale)
            prev = float(bv.importo_previsto or 0)
            rend = float(bv.importo_rendicontato or 0)
            tot_prev += prev; tot_rend += rend
            pct = f"{round(rend / prev * 100, 1)}%" if prev > 0 else "—"
            bv_data.append([label, fmt_eur(prev), fmt_eur(rend),
                            fmt_eur(prev - rend), pct])
        bv_data.append(["TOTALE", fmt_eur(tot_prev), fmt_eur(tot_rend),
                         fmt_eur(tot_prev - tot_rend),
                         f"{round(tot_rend / tot_prev * 100, 1)}%" if tot_prev > 0 else "—"])
        t_bv = tabella(bv_data, [6.5*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2.1*cm], extra_cmds=[
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FB")),
        ])
        story.append(t_bv)

    # Partner
    if partner_rows:
        story.append(Paragraph("Partner del progetto", sezione))
        pa_data = [["Ente", "Tipo", "Paese", "Ruolo", "Budget assegnato"]]
        for pr in partner_rows:
            pa = partner_map.get(str(pr.partner_id))
            pa_data.append([pa.nome if pa else "—", pa.tipo if pa else "—",
                            pa.paese if pa else "—", pr.ruolo,
                            fmt_eur(pr.budget_assegnato) if pr.budget_assegnato else "—"])
        story.append(tabella(pa_data, [5.5*cm, 3*cm, 2*cm, 3*cm, 3.5*cm]))

    # Spese
    if spese:
        story.append(Paragraph("Spese registrate", sezione))
        sp_data = [["Data", "N° documento", "Voce di costo", "Descrizione", "Importo"]]
        tot_sp = 0.0
        for s in spese:
            v = voci_map.get(str(s.voce_id))
            voce_label = Paragraph(f"{v.codice} — {v.descrizione}" if v else "—", normale)
            imp = float(s.importo or 0); tot_sp += imp
            sp_data.append([fmt_d(s.data), s.numero_documento or "—",
                            voce_label, Paragraph(s.descrizione or "—", normale), fmt_eur(imp)])
        sp_data.append(["", "", "", "TOTALE", fmt_eur(tot_sp)])
        t_sp = tabella(sp_data, [2.3*cm, 2.8*cm, 5*cm, 4*cm, 2.9*cm], extra_cmds=[
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FB")),
        ])
        story.append(t_sp)

    # Timesheet
    if timesheet:
        story.append(Paragraph("Timesheet approvati", sezione))
        ts_data = [["Cognome", "Nome", "Mese", "Anno", "Ore totali", "Costo calcolato"]]
        tot_ts = 0.0
        for t in timesheet:
            pe = persone_map.get(str(t.persona_id))
            ore = sum(float(c.ore) for r in t.righe for c in r.celle) if t.righe else 0
            costo = sum(float(c.costo_calcolato or 0) for r in t.righe for c in r.celle) if t.righe else 0
            tot_ts += costo
            ts_data.append([pe.cognome if pe else "—", pe.nome if pe else "—",
                            str(t.mese), str(t.anno), f"{ore:.1f}", fmt_eur(costo)])
        ts_data.append(["", "", "", "", "TOTALE", fmt_eur(tot_ts)])
        t_ts = tabella(ts_data, [3*cm, 3*cm, 1.5*cm, 1.5*cm, 2.5*cm, 5.5*cm], extra_cmds=[
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FB")),
        ])
        story.append(t_ts)

    # Storico SAL
    if sal_list:
        story.append(Paragraph("Storico SAL", sezione))
        sal_data = [["N° SAL", "Inizio periodo", "Fine periodo", "Stato", "Importo erogato"]]
        for s in sal_list:
            sal_data.append([str(s.numero), fmt_d(s.data_inizio), fmt_d(s.data_fine),
                            s.stato, fmt_eur(s.importo_erogato) if s.importo_erogato else "—"])
        story.append(tabella(sal_data, [2*cm, 3.5*cm, 3.5*cm, 3*cm, 5*cm]))

    # Footer
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width=W, color=BLU, thickness=1))
    story.append(Spacer(1, 0.2*cm))
    footer_style = ParagraphStyle("Footer", parent=styles["Normal"],
                                  fontSize=7, textColor=colors.grey, alignment=TA_RIGHT)
    story.append(Paragraph(
        f"Report generato il {datetime.today().strftime('%d/%m/%Y %H:%M')} — {p.acronimo} — Università Pegaso",
        footer_style
    ))

    doc.build(story)
    buf.seek(0)
    nome_pdf = p.acronimo or p.codice or (p.titolo[:30] if p.titolo else "Progetto")
    nome = f"Report_{nome_pdf}_{datetime.today().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{nome}"'})
