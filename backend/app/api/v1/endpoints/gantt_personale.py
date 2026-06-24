# backend/app/api/v1/endpoints/gantt_personale.py
import calendar
from datetime import date
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import or_
from app.core.database import get_db
from app.core.deps import tutti_i_ruoli
from app.models.progetto import Progetto
from app.models.personale import Allocazione, CostoOrarioPersona
from app.models.persona import Persona
from app.models.budget import BudgetVoce, VoceDiCosto
from app.models.timesheet import TimesheetTestata, TimesheetRiga, TimesheetCella
from app.models.budget import Sal

router = APIRouter()


def _mesi_progetto(data_inizio: date, data_fine: date) -> list[dict]:
    """Genera lista mesi M1, M2, ... dalla durata del progetto."""
    mesi = []
    current = date(data_inizio.year, data_inizio.month, 1)
    fine_mese_progetto = date(data_fine.year, data_fine.month, 1)
    num = 1
    while current <= fine_mese_progetto:
        ultimo = calendar.monthrange(current.year, current.month)[1]
        mesi.append({
            "label": f"M{num}",
            "anno": current.year,
            "mese": current.month,
            "data_inizio": current,
            "data_fine": date(current.year, current.month, ultimo),
        })
        num += 1
        current = date(current.year + (current.month // 12), (current.month % 12) + 1, 1)
    return mesi


def _tariffa_effettiva(persona_id, target_date: date, db: Session) -> float:
    """Restituisce il costo orario effettivo alla data indicata."""
    rate = db.query(CostoOrarioPersona).filter(
        CostoOrarioPersona.persona_id == persona_id,
        CostoOrarioPersona.data_inizio <= target_date,
        or_(
            CostoOrarioPersona.data_fine.is_(None),
            CostoOrarioPersona.data_fine >= target_date,
        ),
    ).order_by(CostoOrarioPersona.data_inizio.desc()).first()
    return float(rate.costo_orario) if rate else 0.0


def _calcola_gantt_personale(progetto: Progetto, db: Session) -> dict:
    mesi = _mesi_progetto(progetto.data_inizio, progetto.data_fine)
    num_mesi = len(mesi)

    # ── Pianificazione totale (Budget voce A.1) ──────────────────────────────
    voce_personale = db.query(VoceDiCosto).filter(VoceDiCosto.codice == "A.1").first()
    importo_previsto_personale = 0.0
    if voce_personale:
        bv = db.query(BudgetVoce).filter(
            BudgetVoce.progetto_id == progetto.id,
            BudgetVoce.voce_id == voce_personale.id,
        ).first()
        if bv:
            importo_previsto_personale = float(bv.importo_previsto)

    costo_pianificato_mensile = round(importo_previsto_personale / num_mesi, 2) if num_mesi > 0 else 0.0

    pianificazione_iniziale = {
        "importo_previsto": importo_previsto_personale,
        "per_mese": costo_pianificato_mensile,
        "mesi": [{"label": m["label"], "costo": costo_pianificato_mensile} for m in mesi],
    }

    # ── Timesheet approvati rendicontati per progetto ────────────────────────
    # {(persona_id, anno, mese): {"ore": float, "costo": float}}
    sal_records_rend = db.query(Sal).filter(
        Sal.progetto_id == progetto.id,
        Sal.stato == "rendicontato",
    ).all()
    sal_rendicontati = {str(s.id) for s in sal_records_rend}

    # Mesi (anno, mese) coperti da almeno un SAL rendicontato
    sal_mesi_coperti: set[tuple[int, int]] = set()
    for s in sal_records_rend:
        y, mo = s.data_inizio.year, s.data_inizio.month
        while (y, mo) <= (s.data_fine.year, s.data_fine.month):
            sal_mesi_coperti.add((y, mo))
            mo += 1
            if mo > 12:
                mo = 1
                y += 1

    ts_approvati = db.query(TimesheetTestata).filter(
        TimesheetTestata.progetto_id == progetto.id,
        TimesheetTestata.stato == "approvato",
    ).all()

    dati_reali: dict[tuple, dict] = {}
    for ts in ts_approvati:
        if ts.sal_id is None or str(ts.sal_id) not in sal_rendicontati:
            continue
        ore_ts = 0.0
        costo_ts = 0.0
        for riga in ts.righe:
            if riga.tipo_riga != "progetto":
                continue
            for cella in riga.celle:
                ore_ts += float(cella.ore or 0)
                costo_ts += float(cella.costo_calcolato or 0)
        key = (str(ts.persona_id), ts.anno, ts.mese)
        dati_reali[key] = {"ore": ore_ts, "costo": costo_ts}

    # ── Persone allocate ─────────────────────────────────────────────────────
    allocazioni = db.query(Allocazione).filter(
        Allocazione.progetto_id == progetto.id,
    ).all()

    # Raggruppa allocazioni per persona (somma ore_assegnate, prendi periodo più ampio)
    persone_map: dict[str, dict] = {}
    for alloc in allocazioni:
        pid = str(alloc.persona_id)
        if pid not in persone_map:
            persone_map[pid] = {
                "persona_id": pid,
                "ore_assegnate": 0.0,
                "data_inizio": alloc.data_inizio,
                "data_fine": alloc.data_fine,
            }
        persone_map[pid]["ore_assegnate"] += float(alloc.ore_assegnate)
        persone_map[pid]["data_inizio"] = min(persone_map[pid]["data_inizio"], alloc.data_inizio)
        persone_map[pid]["data_fine"] = max(persone_map[pid]["data_fine"], alloc.data_fine)

    righe_persone = []
    totali_mese = [0.0] * num_mesi

    for pid, alloc_info in persone_map.items():
        persona = db.query(Persona).filter(Persona.id == pid).first()
        if not persona:
            continue

        alloc_inizio = date(alloc_info["data_inizio"].year, alloc_info["data_inizio"].month, 1)
        alloc_fine = date(alloc_info["data_fine"].year, alloc_info["data_fine"].month, 1)

        # Ore già rendicontate per questa persona
        ore_rendicontate = sum(
            dati_reali[(pid, m["anno"], m["mese"])]["ore"]
            for m in mesi
            if (pid, m["anno"], m["mese"]) in dati_reali
              and alloc_inizio <= m["data_inizio"] <= alloc_fine
        )

        # Mesi rimanenti (nell'allocazione, non ancora rendicontati)
        mesi_rimanenti = [
            m for m in mesi
            if alloc_inizio <= m["data_inizio"] <= alloc_fine
            and (pid, m["anno"], m["mese"]) not in dati_reali
            and (m["anno"], m["mese"]) not in sal_mesi_coperti
        ]
        ore_residue = max(0.0, alloc_info["ore_assegnate"] - ore_rendicontate)
        ore_per_mese = round(ore_residue / len(mesi_rimanenti), 2) if mesi_rimanenti else 0.0

        mesi_persona = []
        totale_ore = 0.0
        totale_costo = 0.0

        for idx, m in enumerate(mesi):
            m_inizio = m["data_inizio"]
            in_alloc = alloc_inizio <= m_inizio <= alloc_fine

            if not in_alloc:
                mesi_persona.append({
                    "label": m["label"],
                    "ore": None,
                    "costo": None,
                    "rendicontato": False,
                })
                continue

            key = (pid, m["anno"], m["mese"])
            if key in dati_reali:
                ore = round(dati_reali[key]["ore"], 2)
                costo = round(dati_reali[key]["costo"], 2)
                rendicontato = True
            elif (m["anno"], m["mese"]) in sal_mesi_coperti:
                # Mese coperto da SAL rendicontato ma senza timesheet: 0h/0€ chiuso
                ore = 0.0
                costo = 0.0
                rendicontato = True
            else:
                tariffa = _tariffa_effettiva(pid, m_inizio, db)
                ore = ore_per_mese
                costo = round(ore_per_mese * tariffa, 2)
                rendicontato = False

            totale_ore += ore
            totale_costo += costo
            totali_mese[idx] += costo

            mesi_persona.append({
                "label": m["label"],
                "ore": ore,
                "costo": costo,
                "rendicontato": rendicontato,
            })

        # Tariffa corrente per display
        tariffa_corrente = _tariffa_effettiva(pid, date.today(), db)

        righe_persone.append({
            "persona_id": pid,
            "nome": persona.nome,
            "cognome": persona.cognome,
            "tariffa_corrente": tariffa_corrente,
            "ore_assegnate": alloc_info["ore_assegnate"],
            "alloc_inizio": alloc_info["data_inizio"].isoformat(),
            "alloc_fine": alloc_info["data_fine"].isoformat(),
            "mesi": mesi_persona,
            "totale_ore": round(totale_ore, 2),
            "totale_costo": round(totale_costo, 2),
        })

    # ── Pianificazione Corrente ───────────────────────────────────────────────
    idx_rendicontati = {
        idx for idx, m in enumerate(mesi)
        if any((pid, m["anno"], m["mese"]) in dati_reali for pid in persone_map)
        or (m["anno"], m["mese"]) in sal_mesi_coperti
    }

    pianificazione_corrente_mesi = []

    if not idx_rendicontati:
        # Nessuna spesa reale: pianificazione corrente = pianificazione iniziale
        for idx in range(num_mesi):
            pianificazione_corrente_mesi.append({
                "label": mesi[idx]["label"],
                "costo": costo_pianificato_mensile,
                "rendicontato": False,
            })
    else:
        first_real_idx = min(idx_rendicontati)
        # Budget "consumato" dai mesi precedenti alla prima spesa reale (a quota pianificata)
        budget_mesi_precedenti = costo_pianificato_mensile * first_real_idx
        costo_rendicontato = sum(totali_mese[i] for i in idx_rendicontati)
        budget_residuo = importo_previsto_personale - budget_mesi_precedenti - costo_rendicontato
        # Mesi dopo il primo reale non ancora rendicontati → ricevono la quota residua
        idx_futuri_non_rend = [
            i for i in range(first_real_idx + 1, num_mesi)
            if i not in idx_rendicontati
        ]
        quota_residua = round(budget_residuo / len(idx_futuri_non_rend), 2) if idx_futuri_non_rend else 0.0

        for idx in range(num_mesi):
            if idx < first_real_idx:
                pianificazione_corrente_mesi.append({
                    "label": mesi[idx]["label"],
                    "costo": costo_pianificato_mensile,
                    "rendicontato": False,
                })
            elif idx in idx_rendicontati:
                pianificazione_corrente_mesi.append({
                    "label": mesi[idx]["label"],
                    "costo": round(totali_mese[idx], 2),
                    "rendicontato": True,
                })
            else:
                pianificazione_corrente_mesi.append({
                    "label": mesi[idx]["label"],
                    "costo": quota_residua,
                    "rendicontato": False,
                })

    return {
        "mesi": [{"label": m["label"], "anno": m["anno"], "mese": m["mese"]} for m in mesi],
        "num_mesi": num_mesi,
        "pianificazione_iniziale": pianificazione_iniziale,
        "pianificazione_corrente": {
            "mesi": pianificazione_corrente_mesi,
            "totale": round(importo_previsto_personale, 2),
        },
        "persone": righe_persone,
        "totali_mese": [round(m["costo"], 2) for m in pianificazione_corrente_mesi],
        "totale_complessivo": round(importo_previsto_personale, 2),
    }


def _get_progetto_or_404(progetto_id: str, db: Session) -> Progetto:
    progetto = db.query(Progetto).filter(Progetto.id == progetto_id).first()
    if not progetto:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Progetto non trovato"}})
    return progetto


@router.get("/progetti/{progetto_id}/gantt-personale")
def gantt_personale(
    progetto_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    progetto = _get_progetto_or_404(progetto_id, db)
    return {"data": _calcola_gantt_personale(progetto, db)}


@router.get("/progetti/{progetto_id}/gantt-personale/export/xlsx")
def export_gantt_personale_xlsx(
    progetto_id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    import io

    progetto = _get_progetto_or_404(progetto_id, db)
    data = _calcola_gantt_personale(progetto, db)
    mesi = data["mesi"]
    num_mesi = data["num_mesi"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Personale"

    blu = "185FA5"
    viola = "722ED1"
    verde = "F6FFED"
    verde_testo = "389E0D"
    grigio = "F5F5F5"
    euro = '#,##0.00" €"'

    b = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))
    hfill = PatternFill("solid", fgColor=blu)
    sf = Font(bold=True, size=10)
    sfill = PatternFill("solid", fgColor=grigio)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def st(cell, font=None, fill=None, align=None, number_format=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        if number_format: cell.number_format = number_format
        cell.border = b

    ncols = 2 + num_mesi
    r = 1

    # Titolo
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    titolo = f"GANTT PERSONALE — {progetto.acronimo or progetto.codice} — {progetto.titolo}"
    st(ws.cell(r, 1, titolo), Font(bold=True, color="FFFFFF", size=12), hfill, center)
    ws.row_dimensions[r].height = 26
    r += 1

    # Header
    st(ws.cell(r, 1, "Persona"), sf, sfill, left_al)
    st(ws.cell(r, 2, "Totale"), sf, sfill, center)
    for idx, m in enumerate(mesi):
        st(ws.cell(r, 3 + idx, f"{m['label']}\n{m['mese']}/{m['anno']}"), sf, sfill, center)
    ws.row_dimensions[r].height = 30
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    for idx in range(num_mesi):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 12
    r += 1

    # Pianificazione Iniziale
    pian_iniz = data["pianificazione_iniziale"]
    font_blu = Font(bold=True, color=blu)
    st(ws.cell(r, 1, "Pianificazione Iniziale"), font_blu, align=left_al)
    st(ws.cell(r, 2, pian_iniz["importo_previsto"]), font_blu, align=center, number_format=euro)
    for idx in range(num_mesi):
        st(ws.cell(r, 3 + idx, pian_iniz["per_mese"]), Font(color=blu), align=center, number_format=euro)
    r += 1

    # Pianificazione Corrente
    pian_corr = data["pianificazione_corrente"]
    font_viola = Font(bold=True, color=viola)
    st(ws.cell(r, 1, "Pianificazione Corrente"), font_viola, align=left_al)
    st(ws.cell(r, 2, pian_corr["totale"]), font_viola, align=center, number_format=euro)
    for idx, mc in enumerate(pian_corr["mesi"]):
        if mc["rendicontato"]:
            fill, font = PatternFill("solid", fgColor=verde), Font(bold=True, color=verde_testo)
        else:
            fill, font = None, Font(color=viola)
        st(ws.cell(r, 3 + idx, mc["costo"]), font, fill, center, euro)
    r += 1

    # Persone
    for p in data["persone"]:
        intestazione = f"{p['cognome']} {p['nome']}\n{p['tariffa_corrente']:.2f} €/h · {p['alloc_inizio'][:7]} → {p['alloc_fine'][:7]}"
        st(ws.cell(r, 1, intestazione), align=left_al)
        c_tot = ws.cell(r, 2, p["totale_costo"])
        st(c_tot, Font(bold=True), align=center, number_format=euro)
        c_tot.comment = Comment(f"{p['totale_ore']:.1f} h", "Gestionale Ricerca")
        for idx, md in enumerate(p["mesi"]):
            c = ws.cell(r, 3 + idx)
            if md["ore"] is None:
                st(c, align=center)
                continue
            c.value = md["costo"]
            fill = PatternFill("solid", fgColor=verde) if md["rendicontato"] else None
            st(c, fill=fill, align=center, number_format=euro)
            c.comment = Comment(f"{md['ore']:.1f} h", "Gestionale Ricerca")
        ws.row_dimensions[r].height = 30
        r += 1

    # Totale
    st(ws.cell(r, 1, "TOTALE COSTO MESE"), Font(bold=True), sfill, left_al)
    st(ws.cell(r, 2, data["totale_complessivo"]), Font(bold=True), sfill, center, euro)
    for idx in range(num_mesi):
        st(ws.cell(r, 3 + idx, data["totali_mese"][idx]), Font(bold=True), sfill, center, euro)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"gantt_personale_{progetto.codice}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"}
    )
