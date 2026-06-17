# backend/app/api/v1/endpoints/sal.py
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.deps import tutti_i_ruoli, solo_amministrativo
from app.models.persona import Persona
from app.models.budget import Sal
from app.models.progetto import Progetto
import uuid
from datetime import date

router = APIRouter()

TRANSIZIONI_SAL = {
    "aperto":       {"chiudi": "chiuso"},
    "chiuso":       {"invia": "inviato", "riapri": "aperto"},
    "inviato":      {"contesta": "contestato", "rendiconta": "rendicontato"},
    "contestato":   {"invia": "inviato"},
    "rendicontato": {},
}


def _sal_dict(s: Sal) -> dict:
    return {
        "id": str(s.id),
        "progetto_id": str(s.progetto_id),
        "numero": s.numero,
        "data_inizio": str(s.data_inizio) if s.data_inizio else None,
        "data_fine": str(s.data_fine) if s.data_fine else None,
        "stato": s.stato,
        "importo_tranche": float(s.importo_tranche) if s.importo_tranche else None,
        "importo_erogato": float(s.importo_erogato) if s.importo_erogato else None,
        "data_erogazione": str(s.data_erogazione) if s.data_erogazione else None,
        "data_scadenza_rendiconto": str(s.data_scadenza_rendiconto) if s.data_scadenza_rendiconto else None,
        "motivo_contestazione": s.motivo_contestazione,
    }


def _get_sal_or_404(id: str, db: Session) -> Sal:
    s = db.query(Sal).filter(Sal.id == id).first()
    if not s:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "SAL non trovato"}})
    return s


@router.get("/sal")
def lista_sal(
    progetto_id: str = Query(None),
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    q = db.query(Sal)
    if progetto_id:
        q = q.filter(Sal.progetto_id == progetto_id)
    q = q.order_by(Sal.numero)
    items = q.all()
    return {"data": [_sal_dict(s) for s in items], "meta": {"total": len(items)}}


@router.get("/sal/{id}")
def get_sal(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    s = _get_sal_or_404(id, db)
    return {"data": _sal_dict(s)}


@router.post("/sal")
def crea_sal(
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    progetto_id = body.get("progetto_id")
    if not progetto_id:
        raise HTTPException(status_code=422, detail={"error": {"code": "CAMPO_MANCANTE", "message": "progetto_id obbligatorio"}})

    data_inizio = body.get("data_inizio")
    data_fine = body.get("data_fine")
    if data_inizio and data_fine and str(data_fine) < str(data_inizio):
        raise HTTPException(status_code=422, detail={"error": {"code": "DATE_NON_VALIDE",
            "message": "La data di fine deve essere successiva alla data di inizio"}})

    # Verifica che il progetto esista e sia attivo
    p = db.query(Progetto).filter(Progetto.id == progetto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "Progetto non trovato"}})
    if p.stato != "attivo":
        raise HTTPException(status_code=409, detail={"error": {"code": "PROGETTO_NON_ATTIVO", "message": "Il progetto deve essere in stato attivo per creare un SAL"}})

    # Numero progressivo automatico + controllo continuità
    from datetime import timedelta as _td
    ultimo = db.query(Sal).filter(Sal.progetto_id == progetto_id).order_by(Sal.numero.desc()).first()
    numero = (ultimo.numero + 1) if ultimo else 1

    if ultimo and data_inizio:
        data_inizio_dt = date.fromisoformat(str(data_inizio))
        attesa = ultimo.data_fine + _td(days=1)
        if data_inizio_dt != attesa:
            raise HTTPException(status_code=422, detail={"error": {
                "code": "SAL_NON_CONTINUO",
                "message": (
                    f"La data di inizio deve essere immediatamente successiva alla fine del SAL precedente. "
                    f"Data attesa: {attesa.strftime('%d/%m/%Y')}"
                ),
            }})

    s = Sal(
        id=uuid.uuid4(),
        progetto_id=progetto_id,
        numero=numero,
        data_inizio=body.get("data_inizio"),
        data_fine=body.get("data_fine"),
        importo_tranche=body.get("importo_tranche"),
        data_scadenza_rendiconto=body.get("data_scadenza_rendiconto"),
        stato="aperto",
    )
    db.add(s)
    db.commit()
    db.refresh(s)
    return {"data": _sal_dict(s)}


@router.patch("/sal/{id}")
def aggiorna_sal(
    id: str,
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    s = _get_sal_or_404(id, db)
    if s.stato != "aperto":
        raise HTTPException(status_code=409, detail={"error": {"code": "SAL_NON_MODIFICABILE", "message": "Solo i SAL aperti possono essere modificati"}})
    for k in ("data_inizio", "data_fine", "importo_tranche", "data_scadenza_rendiconto"):
        if k in body:
            setattr(s, k, body[k])
    data_i = str(s.data_inizio) if s.data_inizio else None
    data_f = str(s.data_fine) if s.data_fine else None
    if data_i and data_f and data_f < data_i:
        raise HTTPException(status_code=422, detail={"error": {"code": "DATE_NON_VALIDE",
            "message": "La data di fine deve essere successiva alla data di inizio"}})
    db.commit()
    db.refresh(s)
    return {"data": _sal_dict(s)}


@router.delete("/sal/{id}")
def elimina_sal(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    s = _get_sal_or_404(id, db)
    if s.stato != "aperto":
        raise HTTPException(status_code=409, detail={"error": {"code": "SAL_NON_ELIMINABILE", "message": "Solo i SAL aperti possono essere eliminati"}})
    db.delete(s)
    db.commit()
    return {"data": {"deleted": True}}


@router.post("/sal/{id}/chiudi")
def chiudi_sal(id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    return _transizione_sal(id, "chiudi", db)


@router.post("/sal/{id}/invia")
def invia_sal(id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    return _transizione_sal(id, "invia", db)


@router.post("/sal/{id}/riapri")
def riapri_sal(id: str, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    return _transizione_sal(id, "riapri", db)


@router.post("/sal/{id}/contesta")
def contesta_sal(id: str, body: dict, db: Session = Depends(get_db), utente: Persona = Depends(solo_amministrativo)):
    return _transizione_sal(id, "contesta", db, extra={"motivo_contestazione": body.get("motivo_contestazione")})


@router.post("/sal/{id}/rendiconta")
def rendiconta_sal(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    from app.models.budget import Spesa, BudgetVoce, VoceDiCosto
    from app.models.timesheet import TimesheetTestata, TimesheetCella, TimesheetRiga
    from datetime import date as _date
    import uuid as _uuid

    s = _get_sal_or_404(id, db)

    # Prima cerca timesheet esplicitamente associati al SAL
    timesheet_associati = db.query(TimesheetTestata).filter(
        TimesheetTestata.sal_id == s.id,
        TimesheetTestata.stato == "approvato",
    ).all()

    # Fallback: se nessuno è stato associato esplicitamente,
    # usa tutti gli approvati del progetto nel periodo del SAL
    if not timesheet_associati:
        import calendar as _cal
        tutti = db.query(TimesheetTestata).filter(
            TimesheetTestata.progetto_id == s.progetto_id,
            TimesheetTestata.stato == "approvato",
        ).all()
        for ts in tutti:
            ultimo = _cal.monthrange(ts.anno, ts.mese)[1]
            ts_fine = _date(ts.anno, ts.mese, ultimo)
            ts_inizio = _date(ts.anno, ts.mese, 1)
            # overlap: mese del timesheet si sovrappone al periodo del SAL
            if ts_fine >= s.data_inizio and ts_inizio <= s.data_fine:
                ts.sal_id = s.id
                timesheet_associati.append(ts)

    # Blocca se non ci sono timesheet approvati nel periodo
    if not timesheet_associati:
        raise HTTPException(status_code=422, detail={"error": {
            "code": "NESSUN_TIMESHEET_APPROVATO",
            "message": (
                f"Impossibile rendicontare il SAL: nessun timesheet approvato trovato "
                f"nel periodo {s.data_inizio.strftime('%d/%m/%Y')} — {s.data_fine.strftime('%d/%m/%Y')}. "
                "Approva i timesheet del personale prima di procedere."
            ),
        }})

    costo_personale = 0.0
    for ts in timesheet_associati:
        for riga in ts.righe:
            if riga.tipo_riga != "progetto":
                continue
            for cella in riga.celle:
                costo_personale += float(cella.costo_calcolato or 0)

    # Crea automaticamente la Spesa su "A.1 Personale dipendente" se ci sono costi
    if costo_personale > 0:
        voce_personale = db.query(VoceDiCosto).filter(VoceDiCosto.codice == "A.1").first()
        if voce_personale:
            bv = db.query(BudgetVoce).filter(
                BudgetVoce.progetto_id == s.progetto_id,
                BudgetVoce.voce_id == voce_personale.id,
            ).first()
            spesa_personale = Spesa(
                id=_uuid.uuid4(),
                progetto_id=s.progetto_id,
                voce_id=voce_personale.id,
                sal_id=s.id,
                importo=costo_personale,
                data=s.data_fine,
                descrizione=f"Personale dipendente — SAL #{s.numero}",
                stato="registrata",
            )
            db.add(spesa_personale)

    return _transizione_sal(id, "rendiconta", db)


def _transizione_sal(id: str, azione: str, db: Session, extra: dict = {}):
    s = _get_sal_or_404(id, db)
    nuovo_stato = TRANSIZIONI_SAL.get(s.stato, {}).get(azione)
    if not nuovo_stato:
        raise HTTPException(
            status_code=409,
            detail={"error": {"code": "TRANSIZIONE_NON_VALIDA",
                               "message": f"Impossibile eseguire '{azione}' da stato '{s.stato}'"}},
        )
    s.stato = nuovo_stato
    for k, v in extra.items():
        setattr(s, k, v)
    db.commit()
    db.refresh(s)
    return {"data": _sal_dict(s)}


@router.get("/sal/{id}/dettaglio")
def dettaglio_sal(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.budget import Spesa, BudgetVoce, VoceDiCosto
    from app.models.timesheet import TimesheetTestata, TimesheetRiga, TimesheetCella
    from app.models.personale import CostoOrarioPersona
    from app.models.struttura import WorkPackage
    from sqlalchemy import and_
    from datetime import date

    s = _get_sal_or_404(id, db)

    # ── Spese manuali nel periodo ─────────────────────────────────────────────
    spese = db.query(Spesa).filter(
        and_(
            Spesa.progetto_id == s.progetto_id,
            Spesa.stato == "registrata",
            Spesa.data >= s.data_inizio,
            Spesa.data <= s.data_fine,
        )
    ).all()

    # Spese già associate a questo SAL (selezionate esplicitamente)
    spese_associate = db.query(Spesa).filter(
        Spesa.sal_id == s.id,
        Spesa.stato == "registrata",
    ).all()
    ids_associate = {str(sp.id) for sp in spese_associate}

    spese_list = []
    for sp in spese:
        voce = db.query(VoceDiCosto).filter(VoceDiCosto.id == sp.voce_id).first()
        spese_list.append({
            "id": str(sp.id),
            "tipo": "spesa",
            "data": str(sp.data),
            "descrizione": sp.descrizione or sp.numero_documento or "Spesa",
            "importo": float(sp.importo),
            "voce_id": str(sp.voce_id) if sp.voce_id else None,
            "voce_descrizione": f"{voce.codice} — {voce.descrizione}" if voce else "—",
            "wp_id": None,
            "wp_descrizione": None,
            "selezionato": str(sp.id) in ids_associate or sp.sal_id is None,
            "sal_id": str(sp.sal_id) if sp.sal_id else None,
        })

    # ── Timesheet approvati nel periodo ───────────────────────────────────────
    # Solo persone con allocazione sul progetto
    from app.models.personale import Allocazione
    persone_allocate_sq = db.query(Allocazione.persona_id).filter(
        Allocazione.progetto_id == s.progetto_id
    ).subquery()

    ts_list_raw = db.query(TimesheetTestata).filter(
        and_(
            TimesheetTestata.progetto_id == s.progetto_id,
            TimesheetTestata.stato == "approvato",
            TimesheetTestata.persona_id.in_(persone_allocate_sq),
        )
    ).all()

    timesheet_list = []
    for ts in ts_list_raw:
        # Calcola se il mese del timesheet ricade nel periodo del SAL
        ts_data_inizio = date(ts.anno, ts.mese, 1)
        import calendar
        ultimo_giorno = calendar.monthrange(ts.anno, ts.mese)[1]
        ts_data_fine = date(ts.anno, ts.mese, ultimo_giorno)

        if ts_data_fine < s.data_inizio or ts_data_inizio > s.data_fine:
            continue

        # Costo totale del timesheet (snapshot all'approvazione)
        ore_totali = 0
        costo_totale = 0
        for riga in ts.righe:
            if riga.tipo_riga != "progetto":
                continue
            for cella in riga.celle:
                ore_totali += float(cella.ore)
                costo_totale += float(cella.costo_calcolato or 0)

        if ore_totali == 0:
            continue

        # Persona
        from app.models.persona import Persona as PersonaModel
        persona = db.query(PersonaModel).filter(PersonaModel.id == ts.persona_id).first()
        nome_persona = f"{persona.nome} {persona.cognome}" if persona else "—"

        # Voce di costo personale
        voce_personale = db.query(VoceDiCosto).filter(
            VoceDiCosto.categoria == "personale"
        ).first()

        timesheet_list.append({
            "id": str(ts.id),
            "tipo": "timesheet",
            "data": str(ts_data_inizio),
            "descrizione": f"Timesheet {nome_persona} — {ts.mese}/{ts.anno}",
            "persona_nome": nome_persona,
            "mese": ts.mese,
            "anno": ts.anno,
            "importo": costo_totale,
            "ore": ore_totali,
            "voce_id": str(voce_personale.id) if voce_personale else None,
            "voce_descrizione": f"{voce_personale.codice} — {voce_personale.descrizione}" if voce_personale else "Personale",
            "wp_id": None,
            "wp_descrizione": None,
            "selezionato": str(ts.id) in {str(t) for t in (s.timesheet_ids if hasattr(s, 'timesheet_ids') else [])},
            "sal_id": str(ts.sal_id) if ts.sal_id else None,
        })

    # ── Totali per voce di costo ──────────────────────────────────────────────
    tutti = spese_list + timesheet_list
    totali_per_voce: dict = {}
    for item in tutti:
        if not item["selezionato"]:
            continue
        key = item["voce_id"] or "altro"
        if key not in totali_per_voce:
            totali_per_voce[key] = {
                "voce_id": key,
                "voce_descrizione": item["voce_descrizione"],
                "importo": 0,
            }
        totali_per_voce[key]["importo"] += item["importo"]

    return {
        "data": {
            "sal": _sal_dict(s),
            "voci": spese_list + timesheet_list,
            "totali_per_voce": list(totali_per_voce.values()),
            "totale": sum(i["importo"] for i in tutti if i["selezionato"]),
        }
    }


@router.post("/sal/{id}/associa-voci")
def associa_voci_sal(
    id: str,
    body: dict,
    db: Session = Depends(get_db),
    utente: Persona = Depends(solo_amministrativo),
):
    """Associa/disassocia spese e timesheet a un SAL."""
    from app.models.budget import Spesa
    from app.models.timesheet import TimesheetTestata

    s = _get_sal_or_404(id, db)
    if s.stato != "aperto":
        raise HTTPException(status_code=409, detail={"error": {
            "code": "SAL_NON_MODIFICABILE",
            "message": "Solo i SAL aperti possono essere modificati"
        }})

    spese_ids = body.get("spese_ids", [])
    timesheet_ids = body.get("timesheet_ids", [])

    # Disassocia tutte le spese precedentemente associate a questo SAL
    db.query(Spesa).filter(
        Spesa.sal_id == s.id,
        Spesa.progetto_id == s.progetto_id,
    ).update({"sal_id": None})

    # Associa le spese selezionate
    for spesa_id in spese_ids:
        db.query(Spesa).filter(Spesa.id == spesa_id).update({"sal_id": str(s.id)})

    # Disassocia tutti i timesheet precedentemente associati
    db.query(TimesheetTestata).filter(
        TimesheetTestata.sal_id == s.id,
        TimesheetTestata.progetto_id == s.progetto_id,
    ).update({"sal_id": None})

    # Associa i timesheet selezionati
    for ts_id in timesheet_ids:
        db.query(TimesheetTestata).filter(TimesheetTestata.id == ts_id).update({"sal_id": str(s.id)})

    db.commit()
    return {"data": {"spese_associate": len(spese_ids), "timesheet_associati": len(timesheet_ids)}}


@router.get("/sal/{id}/export/xlsx")
def export_sal_xlsx(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    s = db.query(Sal).filter(Sal.id == id).first()
    if not s:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "SAL non trovato"}})
    if utente.ruolo == "ricercatore":
        raise HTTPException(status_code=403, detail={"error": {
            "code": "FORBIDDEN", "message": "I ricercatori non possono esportare il SAL"}})
    if utente.ruolo == "pi":
        from app.models.personale import Allocazione as _Alloc
        e_pi = db.query(_Alloc).filter(
            _Alloc.persona_id == utente.id,
            _Alloc.progetto_id == s.progetto_id,
            _Alloc.is_pi == True,
        ).first()
        if not e_pi:
            raise HTTPException(status_code=403, detail={"error": {
                "code": "FORBIDDEN", "message": "Solo il PI del progetto può esportare il SAL"}})
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.budget import Spesa, VoceDiCosto
    from app.models.timesheet import TimesheetTestata
    from app.models.persona import Persona as PersonaModel
    from app.models.progetto import Progetto
    from sqlalchemy import and_
    import io, calendar

    progetto = db.query(Progetto).filter(Progetto.id == s.progetto_id).first()

    wb = Workbook()
    ws = wb.active
    ws.title = f"SAL {s.numero}"

    blu = "185FA5"
    grigio = "F5F5F5"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=blu)
    sub_font = Font(bold=True, size=10)
    sub_fill = PatternFill("solid", fgColor=grigio)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def style(cell, font=None, fill=None, align=None):
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        cell.border = border

    riga = 1

    # ── Intestazione ──
    ws.merge_cells(f"A{riga}:F{riga}")
    c = ws.cell(riga, 1, f"SAL {s.numero} — {progetto.acronimo or progetto.codice if progetto else ''}")
    style(c, header_font, header_fill, center)
    ws.row_dimensions[riga].height = 28
    riga += 1

    info = [
        ("Progetto", progetto.titolo if progetto else "—"),
        ("Periodo", f"{s.data_inizio.strftime('%d/%m/%Y')} → {s.data_fine.strftime('%d/%m/%Y')}"),
        ("Stato", s.stato.upper()),
        ("Scad. rendiconto", s.data_scadenza_rendiconto.strftime('%d/%m/%Y') if s.data_scadenza_rendiconto else "—"),
    ]
    for label, val in info:
        ws.cell(riga, 1, label).font = Font(bold=True)
        ws.cell(riga, 1).border = border
        ws.merge_cells(f"B{riga}:F{riga}")
        style(ws.cell(riga, 2, val), align=left)
        riga += 1

    riga += 1

    # ── Spese ──
    spese = db.query(Spesa).filter(
        and_(Spesa.progetto_id == s.progetto_id, Spesa.sal_id == s.id, Spesa.stato == "registrata")
    ).all()

    if spese:
        ws.merge_cells(f"A{riga}:F{riga}")
        style(ws.cell(riga, 1, "SPESE"), sub_font, sub_fill, center)
        riga += 1

        headers_spese = ["Data", "N° Documento", "Voce di costo", "Descrizione", "Importo"]
        for col, h in enumerate(headers_spese, 1):
            style(ws.cell(riga, col, h), sub_font, sub_fill, center)
        riga += 1

        tot_spese = 0
        for sp in spese:
            voce = db.query(VoceDiCosto).filter(VoceDiCosto.id == sp.voce_id).first()
            row_data = [
                str(sp.data) if sp.data else "—",
                sp.numero_documento or "—",
                f"{voce.codice} — {voce.descrizione}" if voce else "—",
                sp.descrizione or "—",
                float(sp.importo),
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(riga, col, val)
                style(c, align=right if col == 5 else left)
            tot_spese += float(sp.importo)
            riga += 1

        ws.cell(riga, 4, "Totale spese").font = Font(bold=True)
        ws.cell(riga, 4).border = border
        c = ws.cell(riga, 5, tot_spese)
        style(c, Font(bold=True), align=right)
        riga += 2

    # ── Timesheet ──
    ts_list = db.query(TimesheetTestata).filter(
        and_(TimesheetTestata.progetto_id == s.progetto_id,
             TimesheetTestata.sal_id == s.id,
             TimesheetTestata.stato == "approvato")
    ).all()

    if ts_list:
        ws.merge_cells(f"A{riga}:F{riga}")
        style(ws.cell(riga, 1, "TIMESHEET"), sub_font, sub_fill, center)
        riga += 1

        headers_ts = ["Persona", "Mese/Anno", "Ore progetto", "Costo orario", "Costo totale"]
        for col, h in enumerate(headers_ts, 1):
            style(ws.cell(riga, col, h), sub_font, sub_fill, center)
        riga += 1

        tot_ts = 0
        for ts in ts_list:
            persona = db.query(PersonaModel).filter(PersonaModel.id == ts.persona_id).first()
            ore = sum(float(c.ore) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
            costo = sum(float(c.costo_calcolato or 0) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
            costo_orario = costo / ore if ore > 0 else 0
            row_data = [
                f"{persona.cognome} {persona.nome}" if persona else "—",
                f"{ts.mese:02d}/{ts.anno}",
                ore, round(costo_orario, 2), round(costo, 2),
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(riga, col, val)
                style(c, align=right if col >= 3 else left)
            tot_ts += costo
            riga += 1

        ws.cell(riga, 4, "Totale personale").font = Font(bold=True)
        ws.cell(riga, 4).border = border
        c = ws.cell(riga, 5, round(tot_ts, 2))
        style(c, Font(bold=True), align=right)
        riga += 2

    # ── Totali per voce ──
    ws.merge_cells(f"A{riga}:F{riga}")
    style(ws.cell(riga, 1, "TOTALI PER VOCE DI COSTO"), sub_font, sub_fill, center)
    riga += 1

    totali_voce: dict = {}
    for sp in spese:
        voce = db.query(VoceDiCosto).filter(VoceDiCosto.id == sp.voce_id).first()
        key = voce.descrizione if voce else "Altro"
        totali_voce[key] = totali_voce.get(key, 0) + float(sp.importo)
    for ts in ts_list:
        costo = sum(float(c.costo_calcolato or 0) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
        totali_voce["Personale dipendente"] = totali_voce.get("Personale dipendente", 0) + costo

    totale_generale = 0
    for voce_desc, importo in totali_voce.items():
        ws.cell(riga, 1, voce_desc).border = border
        ws.merge_cells(f"A{riga}:D{riga}")
        c = ws.cell(riga, 5, round(importo, 2))
        style(c, align=right)
        totale_generale += importo
        riga += 1

    ws.cell(riga, 4, "TOTALE GENERALE").font = Font(bold=True, color="FFFFFF")
    ws.cell(riga, 4).fill = PatternFill("solid", fgColor=blu)
    ws.cell(riga, 4).border = border
    c = ws.cell(riga, 5, round(totale_generale, 2))
    style(c, Font(bold=True), PatternFill("solid", fgColor=blu), right)
    c.font = Font(bold=True, color="FFFFFF")

    # Larghezze colonne
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import os as _os
    from app.services.storage import progetto_dir, _safe
    _codice = progetto.codice if progetto else "export"
    _sal_stem = f"{s.numero}_{_safe(_codice)}_{date.today().strftime('%d%m%Y')}"
    _output_dir = progetto_dir(_codice, "sal", _sal_stem)
    _os.makedirs(_output_dir, exist_ok=True)
    filename = f"SAL_{_sal_stem}.xlsx"
    _dst = _os.path.join(_output_dir, filename)
    with open(_dst, "wb") as _fh:
        _fh.write(buf.read())
    s.xlsx_path = _dst
    db.commit()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/sal/{id}/export/pdf")
def export_sal_pdf(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable)
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from app.models.budget import Spesa, VoceDiCosto
    from app.models.timesheet import TimesheetTestata
    from app.models.persona import Persona as PersonaModel
    from app.models.personale import Allocazione
    from sqlalchemy import and_
    import io
    from datetime import date

    s = db.query(Sal).filter(Sal.id == id).first()
    if not s:
        raise HTTPException(status_code=404, detail={"error": {"code": "NOT_FOUND", "message": "SAL non trovato"}})

    # Permessi (stessa logica export xlsx)
    if utente.ruolo == "ricercatore":
        raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN", "message": "Accesso non consentito"}})
    if utente.ruolo == "pi":
        e_pi = db.query(Allocazione).filter(
            Allocazione.persona_id == utente.id,
            Allocazione.progetto_id == s.progetto_id,
            Allocazione.is_pi == True,
        ).first()
        if not e_pi:
            raise HTTPException(status_code=403, detail={"error": {"code": "FORBIDDEN", "message": "Solo il PI del progetto può esportare il SAL"}})

    progetto = db.query(Progetto).filter(Progetto.id == s.progetto_id).first()

    # PI e amministrativo
    pi_alloc = db.query(Allocazione).filter(
        Allocazione.progetto_id == s.progetto_id, Allocazione.is_pi == True
    ).first()
    pi_persona = db.query(PersonaModel).filter(PersonaModel.id == pi_alloc.persona_id).first() if pi_alloc else None
    pi_nome = f"{pi_persona.nome} {pi_persona.cognome}" if pi_persona else "—"

    amm_persona = None
    if progetto and progetto.amministrativo_id:
        amm_persona = db.query(PersonaModel).filter(PersonaModel.id == progetto.amministrativo_id).first()
    amm_nome = f"{amm_persona.nome} {amm_persona.cognome}" if amm_persona else "—"

    # Dati spese e timesheet
    spese = db.query(Spesa).filter(
        and_(Spesa.progetto_id == s.progetto_id, Spesa.sal_id == s.id, Spesa.stato == "registrata")
    ).all()
    ts_list = db.query(TimesheetTestata).filter(
        and_(TimesheetTestata.progetto_id == s.progetto_id,
             TimesheetTestata.sal_id == s.id,
             TimesheetTestata.stato == "approvato")
    ).all()

    # ── Stili ──
    BLU = colors.HexColor("#185FA5")
    GRIGIO_CHIARO = colors.HexColor("#F5F7FA")
    GRIGIO_BORDO = colors.HexColor("#D0D5DD")
    BIANCO = colors.white
    NERO = colors.HexColor("#1A1A2E")

    styles = getSampleStyleSheet()
    s_titolo = ParagraphStyle("titolo", fontSize=16, textColor=BIANCO, fontName="Helvetica-Bold",
                               alignment=TA_CENTER, spaceAfter=0)
    s_label = ParagraphStyle("label", fontSize=9, textColor=colors.HexColor("#6B7280"),
                              fontName="Helvetica", leading=13)
    s_valore = ParagraphStyle("valore", fontSize=10, textColor=NERO,
                               fontName="Helvetica-Bold", leading=13)
    s_section = ParagraphStyle("section", fontSize=11, textColor=BLU,
                                fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    s_body = ParagraphStyle("body", fontSize=9, fontName="Helvetica", leading=12)
    s_footer = ParagraphStyle("footer", fontSize=8, textColor=colors.HexColor("#9CA3AF"),
                               fontName="Helvetica", alignment=TA_CENTER)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)

    W = A4[0] - 4*cm  # larghezza utile

    story = []

    # ── Banner intestazione ──
    banner_data = [[Paragraph(
        f"SAL {s.numero} &nbsp;·&nbsp; {progetto.acronimo or progetto.codice if progetto else ''}",
        s_titolo
    )]]
    banner = Table(banner_data, colWidths=[W])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BLU),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(banner)
    story.append(Spacer(1, 0.4*cm))

    # ── Scheda info progetto ──
    nome_prog = progetto.titolo if progetto else "—"
    periodo = f"{s.data_inizio.strftime('%d/%m/%Y') if s.data_inizio else '—'} → {s.data_fine.strftime('%d/%m/%Y') if s.data_fine else '—'}"
    scad = s.data_scadenza_rendiconto.strftime('%d/%m/%Y') if s.data_scadenza_rendiconto else "—"
    stato_label = s.stato.upper()

    def cella_info(label, valore):
        return [Paragraph(label, s_label), Paragraph(valore, s_valore)]

    info_data = [
        [*cella_info("PROGETTO", nome_prog), *cella_info("PERIODO", periodo)],
        [*cella_info("PI", pi_nome), *cella_info("AMMINISTRATIVO", amm_nome)],
        [*cella_info("STATO", stato_label), *cella_info("SCADENZA RENDICONTO", scad)],
    ]
    col_w = W / 4
    info_table = Table(info_data, colWidths=[col_w * 0.8, col_w * 1.2, col_w * 0.8, col_w * 1.2])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GRIGIO_CHIARO),
        ("BOX", (0, 0), (-1, -1), 0.5, GRIGIO_BORDO),
        ("LINEAFTER", (1, 0), (1, -1), 0.5, GRIGIO_BORDO),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, GRIGIO_BORDO),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    def fmt_eur(v): return f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    def fmt_data(d): return d.strftime('%d/%m/%Y') if d else "—"

    th_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLU),
        ("TEXTCOLOR", (0, 0), (-1, 0), BIANCO),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BIANCO, GRIGIO_CHIARO]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, GRIGIO_BORDO),
        ("BOX", (0, 0), (-1, -1), 0.5, GRIGIO_BORDO),
    ])

    # ── Sezione Spese ──
    if spese:
        story.append(Paragraph("Spese documentate", s_section))
        tot_spese = 0
        righe_spese = [["Data", "N° Documento", "Voce di costo", "Descrizione", "Importo"]]
        for sp in spese:
            voce = db.query(VoceDiCosto).filter(VoceDiCosto.id == sp.voce_id).first()
            importo = float(sp.importo)
            righe_spese.append([
                fmt_data(sp.data),
                sp.numero_documento or "—",
                voce.descrizione if voce else "—",
                Paragraph(sp.descrizione or "—", s_body),
                Paragraph(f'<para alignment="right">{fmt_eur(importo)}</para>', s_body),
            ])
            tot_spese += importo
        righe_spese.append(["", "", "", Paragraph('<b>Totale spese</b>', s_body),
                             Paragraph(f'<para alignment="right"><b>{fmt_eur(tot_spese)}</b></para>', s_body)])

        t_spese = Table(righe_spese, colWidths=[2.2*cm, 2.8*cm, 4.5*cm, None, 2.8*cm])
        t_spese.setStyle(th_style)
        t_spese.setStyle(TableStyle([
            ("BACKGROUND", (0, len(righe_spese)-1), (-1, -1), GRIGIO_CHIARO),
            ("FONTNAME", (0, len(righe_spese)-1), (-1, -1), "Helvetica-Bold"),
        ]))
        story.append(t_spese)
        story.append(Spacer(1, 0.4*cm))

    # ── Sezione Timesheet ──
    if ts_list:
        story.append(Paragraph("Personale — Timesheet approvati", s_section))
        tot_ts = 0
        righe_ts = [["Persona", "Mese/Anno", "Ore progetto", "Costo orario", "Costo totale"]]
        for ts in ts_list:
            persona = db.query(PersonaModel).filter(PersonaModel.id == ts.persona_id).first()
            ore = sum(float(c.ore) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
            costo = sum(float(c.costo_calcolato or 0) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
            costo_orario = costo / ore if ore > 0 else 0
            righe_ts.append([
                f"{persona.cognome} {persona.nome}" if persona else "—",
                f"{ts.mese:02d}/{ts.anno}",
                Paragraph(f'<para alignment="right">{ore:.1f}h</para>', s_body),
                Paragraph(f'<para alignment="right">{fmt_eur(costo_orario)}/h</para>', s_body),
                Paragraph(f'<para alignment="right">{fmt_eur(costo)}</para>', s_body),
            ])
            tot_ts += costo
        righe_ts.append(["", "", "", Paragraph('<b>Totale personale</b>', s_body),
                          Paragraph(f'<para alignment="right"><b>{fmt_eur(tot_ts)}</b></para>', s_body)])

        t_ts = Table(righe_ts, colWidths=[5*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm])
        t_ts.setStyle(th_style)
        t_ts.setStyle(TableStyle([
            ("BACKGROUND", (0, len(righe_ts)-1), (-1, -1), GRIGIO_CHIARO),
            ("FONTNAME", (0, len(righe_ts)-1), (-1, -1), "Helvetica-Bold"),
        ]))
        story.append(t_ts)
        story.append(Spacer(1, 0.4*cm))

    # ── Riepilogo per voce ──
    story.append(Paragraph("Riepilogo per voce di costo", s_section))
    totali_voce: dict = {}
    for sp in spese:
        voce = db.query(VoceDiCosto).filter(VoceDiCosto.id == sp.voce_id).first()
        key = voce.descrizione if voce else "Altro"
        totali_voce[key] = totali_voce.get(key, 0) + float(sp.importo)
    for ts in ts_list:
        costo = sum(float(c.costo_calcolato or 0) for r in ts.righe if r.tipo_riga == "progetto" for c in r.celle)
        if costo > 0:
            totali_voce["Personale dipendente"] = totali_voce.get("Personale dipendente", 0) + costo

    totale_gen = sum(totali_voce.values())
    righe_ric = [["Voce di costo", "Importo"]]
    for desc, imp in totali_voce.items():
        righe_ric.append([desc, Paragraph(f'<para alignment="right">{fmt_eur(imp)}</para>', s_body)])
    righe_ric.append([
        Paragraph("<b>TOTALE GENERALE</b>", s_body),
        Paragraph(f'<para alignment="right"><b>{fmt_eur(totale_gen)}</b></para>', s_body),
    ])

    t_ric = Table(righe_ric, colWidths=[None, 3.5*cm])
    t_ric.setStyle(th_style)
    t_ric.setStyle(TableStyle([
        ("BACKGROUND", (0, len(righe_ric)-1), (-1, -1), BLU),
        ("TEXTCOLOR", (0, len(righe_ric)-1), (-1, -1), BIANCO),
        ("FONTNAME", (0, len(righe_ric)-1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(t_ric)

    # ── Footer ──
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRIGIO_BORDO))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Documento generato il {date.today().strftime('%d/%m/%Y')} — Gestionale Ricerca",
        s_footer
    ))

    doc.build(story)
    buf.seek(0)
    import os as _os
    from app.services.storage import progetto_dir, _safe
    _codice = progetto.codice if progetto else "export"
    _sal_stem = f"{s.numero}_{_safe(_codice)}_{date.today().strftime('%d%m%Y')}"
    _output_dir = progetto_dir(_codice, "sal", _sal_stem)
    _os.makedirs(_output_dir, exist_ok=True)
    nome_file = f"SAL_{_sal_stem}.pdf"
    _dst = _os.path.join(_output_dir, nome_file)
    with open(_dst, "wb") as _fh:
        _fh.write(buf.read())
    s.pdf_path = _dst
    db.commit()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f"attachment; filename={nome_file}"})


@router.get("/notifiche")
def notifiche(
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from datetime import date, timedelta
    from app.models.timesheet import TimesheetTestata
    from app.models.progetto import Progetto
    from app.models.personale import Allocazione

    oggi = date.today()
    tra_30 = oggi + timedelta(days=30)
    risultati = []

    # SAL in scadenza entro 30 giorni
    if utente.ruolo in ("amministrativo", "management", "pi"):
        if utente.ruolo in ("amministrativo", "management"):
            sal_q = db.query(Sal)
        else:
            # PI vede solo i SAL dei progetti a cui partecipa
            proj_ids = [str(a.progetto_id) for a in
                       db.query(Allocazione).filter(Allocazione.persona_id == utente.id).all()]
            sal_q = db.query(Sal).filter(Sal.progetto_id.in_(proj_ids))

        sal_scadenza = sal_q.filter(
            Sal.stato.in_(["aperto", "chiuso"]),
            Sal.data_scadenza_rendiconto != None,
            Sal.data_scadenza_rendiconto >= oggi,
            Sal.data_scadenza_rendiconto <= tra_30,
        ).all()

        for s in sal_scadenza:
            p = db.query(Progetto).filter(Progetto.id == s.progetto_id).first()
            giorni = (s.data_scadenza_rendiconto - oggi).days
            risultati.append({
                "id": str(s.id),
                "tipo": "sal_scadenza",
                "titolo": f"SAL {s.numero} in scadenza",
                "messaggio": f"{p.acronimo or p.codice} — scade tra {giorni} giorn{'o' if giorni == 1 else 'i'}",
                "giorni_rimanenti": giorni,
                "link": f"/sal/{s.id}",
                "urgente": giorni <= 7,
            })

    # Timesheet in attesa di approvazione
    if utente.ruolo in ("pi", "amministrativo"):
        ts_pendenti = db.query(TimesheetTestata).filter(
            TimesheetTestata.stato == "inviato"
        ).all()

        for ts in ts_pendenti:
            p = db.query(Progetto).filter(Progetto.id == ts.progetto_id).first()
            from app.models.persona import Persona as PersonaModel
            persona = db.query(PersonaModel).filter(PersonaModel.id == ts.persona_id).first()
            risultati.append({
                "id": str(ts.id),
                "tipo": "timesheet_pendente",
                "titolo": "Timesheet da approvare",
                "messaggio": f"{persona.nome} {persona.cognome} — {ts.mese:02d}/{ts.anno} ({p.acronimo or p.codice if p else ''})",
                "giorni_rimanenti": None,
                "link": f"/timesheet/{ts.id}",
                "urgente": False,
            })

    # Ordina: urgenti prima, poi per giorni rimanenti
    risultati.sort(key=lambda x: (not x["urgente"], x["giorni_rimanenti"] if x["giorni_rimanenti"] is not None else 999))

    # Aggiungi notifiche personali dal sistema
    try:
        from app.models.notifica import Notifica
        notifiche_db = db.query(Notifica).filter(
            Notifica.persona_id == utente.id,
            Notifica.letta == False,
        ).order_by(Notifica.created_at.desc()).all()

        for n in notifiche_db:
            risultati.insert(0, {
                "id": str(n.id),
                "tipo": n.tipo,
                "titolo": n.titolo,
                "messaggio": n.messaggio or "",
                "giorni_rimanenti": None,
                "link": n.link or "/",
                "urgente": False,
            })
    except Exception:
        pass

    return {"data": risultati, "meta": {"totale": len(risultati)}}


@router.post("/notifiche/{id}/letta")
def segna_notifica_letta(
    id: str,
    db: Session = Depends(get_db),
    utente: Persona = Depends(tutti_i_ruoli),
):
    from app.models.notifica import Notifica
    n = db.query(Notifica).filter(Notifica.id == id, Notifica.persona_id == utente.id).first()
    if n:
        n.letta = True
        db.commit()
    return {"data": {"ok": True}}
